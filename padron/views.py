from django.shortcuts import render, redirect, get_object_or_404
from django.http.response import JsonResponse
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.db import IntegrityError
from django.contrib.auth.decorators import login_required

from .models import Padron, Provincia, Distrito, Compromiso

# report excel
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from datetime import datetime

# tablas por redes
from django.db.models import Sum, F, FloatField, ExpressionWrapper,Count, Value, Q
from django.db.models.functions import Cast, Round,Coalesce 

################################################
# SITUACION PADRON NOMINAL - VISITA DOMICILARIO
################################################
@login_required
def index(request):
    p_chyo = 'CHANCHAMAYO'
    p_chupaca = 'CHUPACA'
    p_concepcion = 'CONCEPCION'
    p_huancayo = 'HUANCAYO'
    p_jauja = 'JAUJA'
    p_junin = 'JUNIN'
    p_satipo = 'SATIPO'
    p_tarma = 'TARMA'
    p_yauli = 'YAULI'
       
    t_chyo = Padron.objects.values('distrito').annotate( 
                                                        sum_0_28d =Sum('edad_0_28_dias'),
                                                        sum_0_5m =Sum('edad_0_5_meses'),
                                                        sum_6_11m=Sum('edad_6_11_meses'),
                                                        sum_12m=Sum('menores_de_12_meses'),
                                                        sum_1a=Sum('edad_1_anio'),
                                                        sum_2a=Sum('edad_2_anios'),
                                                        sum_3a=Sum('edad_3_anios'),
                                                        sum_4a=Sum('edad_4_anios'),
                                                        sum_5a=Sum('edad_5_anios'),
                                                        sum_num = Count('num_num_doc', filter=Q(num_num_doc=0)),
                                                        sum_seguro= Count('num_seguro',filter=Q(num_seguro=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0)),
                                                        sum_frecuencia= Count('num_frecuencia', filter=Q(num_frecuencia=0)),
                                                        sum_entidad= Count('num_entidad_reniec', filter=Q(num_entidad_reniec=1)),
                                                        ).filter(provincia=p_chyo).order_by('distrito')                                                                                   

    t_chupaca = Padron.objects.values('distrito').annotate( 
                                                        sum_0_28d =Sum('edad_0_28_dias'),
                                                        sum_0_5m =Sum('edad_0_5_meses'),
                                                        sum_6_11m=Sum('edad_6_11_meses'),
                                                        sum_12m=Sum('menores_de_12_meses'),
                                                        sum_1a=Sum('edad_1_anio'),
                                                        sum_2a=Sum('edad_2_anios'),
                                                        sum_3a=Sum('edad_3_anios'),
                                                        sum_4a=Sum('edad_4_anios'),
                                                        sum_5a=Sum('edad_5_anios'),
                                                        sum_num = Count('num_num_doc', filter=Q(num_num_doc=0)),
                                                        sum_seguro= Count('num_seguro',filter=Q(num_seguro=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0)),
                                                        sum_frecuencia= Count('num_frecuencia', filter=Q(num_frecuencia=0)),
                                                        sum_entidad= Count('num_entidad_reniec', filter=Q(num_entidad_reniec=1)),
                                                        ).filter(provincia=p_chupaca).order_by('distrito')                                          
    
    t_concepcion = Padron.objects.values('distrito').annotate( 
                                                        sum_0_28d =Sum('edad_0_28_dias'),
                                                        sum_0_5m =Sum('edad_0_5_meses'),
                                                        sum_6_11m=Sum('edad_6_11_meses'),
                                                        sum_12m=Sum('menores_de_12_meses'),
                                                        sum_1a=Sum('edad_1_anio'),
                                                        sum_2a=Sum('edad_2_anios'),
                                                        sum_3a=Sum('edad_3_anios'),
                                                        sum_4a=Sum('edad_4_anios'),
                                                        sum_5a=Sum('edad_5_anios'),
                                                        sum_num = Count('num_num_doc', filter=Q(num_num_doc=0)),
                                                        sum_seguro= Count('num_seguro',filter=Q(num_seguro=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0)),
                                                        sum_frecuencia= Count('num_frecuencia', filter=Q(num_frecuencia=0)),
                                                        sum_entidad= Count('num_entidad_reniec', filter=Q(num_entidad_reniec=1)),
                                                        ).filter(provincia=p_concepcion).order_by('distrito')                                                        
    
    t_huancayo = Padron.objects.values('distrito').annotate( 
                                                        sum_0_28d =Sum('edad_0_28_dias'),
                                                        sum_0_5m =Sum('edad_0_5_meses'),
                                                        sum_6_11m=Sum('edad_6_11_meses'),
                                                        sum_12m=Sum('menores_de_12_meses'),
                                                        sum_1a=Sum('edad_1_anio'),
                                                        sum_2a=Sum('edad_2_anios'),
                                                        sum_3a=Sum('edad_3_anios'),
                                                        sum_4a=Sum('edad_4_anios'),
                                                        sum_5a=Sum('edad_5_anios'),
                                                        sum_num = Count('num_num_doc', filter=Q(num_num_doc=0)),
                                                        sum_seguro= Count('num_seguro',filter=Q(num_seguro=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0)),
                                                        sum_frecuencia= Count('num_frecuencia', filter=Q(num_frecuencia=0)),
                                                        sum_entidad= Count('num_entidad_reniec', filter=Q(num_entidad_reniec=1)),
                                                        ).filter(provincia=p_huancayo).order_by('distrito')      
    
    t_jauja = Padron.objects.values('distrito').annotate( 
                                                        sum_0_28d =Sum('edad_0_28_dias'),
                                                        sum_0_5m =Sum('edad_0_5_meses'),
                                                        sum_6_11m=Sum('edad_6_11_meses'),
                                                        sum_12m=Sum('menores_de_12_meses'),
                                                        sum_1a=Sum('edad_1_anio'),
                                                        sum_2a=Sum('edad_2_anios'),
                                                        sum_3a=Sum('edad_3_anios'),
                                                        sum_4a=Sum('edad_4_anios'),
                                                        sum_5a=Sum('edad_5_anios'),
                                                        sum_num = Count('num_num_doc', filter=Q(num_num_doc=0)),
                                                        sum_seguro= Count('num_seguro',filter=Q(num_seguro=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0)),
                                                        sum_frecuencia= Count('num_frecuencia', filter=Q(num_frecuencia=0)),
                                                        sum_entidad= Count('num_entidad_reniec', filter=Q(num_entidad_reniec=1)),
                                                        ).filter(provincia=p_jauja).order_by('distrito')      
    
    t_junin = Padron.objects.values('distrito').annotate( 
                                                        sum_0_28d =Sum('edad_0_28_dias'),
                                                        sum_0_5m =Sum('edad_0_5_meses'),
                                                        sum_6_11m=Sum('edad_6_11_meses'),
                                                        sum_12m=Sum('menores_de_12_meses'),
                                                        sum_1a=Sum('edad_1_anio'),
                                                        sum_2a=Sum('edad_2_anios'),
                                                        sum_3a=Sum('edad_3_anios'),
                                                        sum_4a=Sum('edad_4_anios'),
                                                        sum_5a=Sum('edad_5_anios'),
                                                        sum_num = Count('num_num_doc', filter=Q(num_num_doc=0)),
                                                        sum_seguro= Count('num_seguro',filter=Q(num_seguro=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0)),
                                                        sum_frecuencia= Count('num_frecuencia', filter=Q(num_frecuencia=0)),
                                                        sum_entidad= Count('num_entidad_reniec', filter=Q(num_entidad_reniec=1)),
                                                        ).filter(provincia=p_junin).order_by('distrito') 
    
    t_satipo = Padron.objects.values('distrito').annotate( 
                                                        sum_0_28d =Sum('edad_0_28_dias'),
                                                        sum_0_5m =Sum('edad_0_5_meses'),
                                                        sum_6_11m=Sum('edad_6_11_meses'),
                                                        sum_12m=Sum('menores_de_12_meses'),
                                                        sum_1a=Sum('edad_1_anio'),
                                                        sum_2a=Sum('edad_2_anios'),
                                                        sum_3a=Sum('edad_3_anios'),
                                                        sum_4a=Sum('edad_4_anios'),
                                                        sum_5a=Sum('edad_5_anios'),
                                                        sum_num = Count('num_num_doc', filter=Q(num_num_doc=0)),
                                                        sum_seguro= Count('num_seguro',filter=Q(num_seguro=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0)),
                                                        sum_frecuencia= Count('num_frecuencia', filter=Q(num_frecuencia=0)),
                                                        sum_entidad= Count('num_entidad_reniec', filter=Q(num_entidad_reniec=1)),
                                                        ).filter(provincia=p_satipo).order_by('distrito')        
    
    t_tarma = Padron.objects.values('distrito').annotate( 
                                                        sum_0_28d =Sum('edad_0_28_dias'),
                                                        sum_0_5m =Sum('edad_0_5_meses'),
                                                        sum_6_11m=Sum('edad_6_11_meses'),
                                                        sum_12m=Sum('menores_de_12_meses'),
                                                        sum_1a=Sum('edad_1_anio'),
                                                        sum_2a=Sum('edad_2_anios'),
                                                        sum_3a=Sum('edad_3_anios'),
                                                        sum_4a=Sum('edad_4_anios'),
                                                        sum_5a=Sum('edad_5_anios'),
                                                        sum_num = Count('num_num_doc', filter=Q(num_num_doc=0)),
                                                        sum_seguro= Count('num_seguro',filter=Q(num_seguro=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0)),
                                                        sum_frecuencia= Count('num_frecuencia', filter=Q(num_frecuencia=0)),
                                                        sum_entidad= Count('num_entidad_reniec', filter=Q(num_entidad_reniec=1)),
                                                        ).filter(provincia=p_tarma).order_by('distrito')  
    
    t_yauli = Padron.objects.values('distrito').annotate( 
                                                        sum_0_28d =Sum('edad_0_28_dias'),
                                                        sum_0_5m =Sum('edad_0_5_meses'),
                                                        sum_6_11m=Sum('edad_6_11_meses'),
                                                        sum_12m=Sum('menores_de_12_meses'),
                                                        sum_1a=Sum('edad_1_anio'),
                                                        sum_2a=Sum('edad_2_anios'),
                                                        sum_3a=Sum('edad_3_anios'),
                                                        sum_4a=Sum('edad_4_anios'),
                                                        sum_5a=Sum('edad_5_anios'),
                                                        sum_num = Count('num_num_doc', filter=Q(num_num_doc=0)),
                                                        sum_seguro= Count('num_seguro',filter=Q(num_seguro=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0)),
                                                        sum_frecuencia= Count('num_frecuencia', filter=Q(num_frecuencia=0)),
                                                        sum_entidad= Count('num_entidad_reniec', filter=Q(num_entidad_reniec=1)),
                                                        ).filter(provincia=p_yauli).order_by('distrito')  
    
    context = {
                't_chyo'      : t_chyo,
                't_chupaca'   : t_chupaca,
                't_concepcion': t_concepcion,
                't_huancayo'  : t_huancayo,
                't_jauja'     : t_jauja,
                't_junin'     : t_junin,
                't_satipo'    : t_satipo,
                't_tarma'     : t_tarma,
                't_yauli'     : t_yauli,
              }
    
    return render(request, 'situacion.html', context)

#####################################################################
#---  GRAFICOS PRINCIPAL CANTIDAD DE NIÑOS ---------------------------
def get_chart_padron_edades(_request):
    # Consulta para obtener los datos de ventas
    
    # Formatear los datos para pasarlos a la plantilla
    colors = ['#5470C6', '#91CC75', '#EE6666'];   
    
    serie=['1446','9275','10063','19338','20874','21079','23129','24515','24554']
   
    chart_padron_edades = {
        'tooltip':{
            'show': True,
            'trigger': "axis",
            'triggerOn': "mousemove|click"    
        },
        'legend':{# Nombre para la leyenda
        },       
        'xAxis':[
            {
                'type':"category",
                'data':["0 a 28 d","0 a 5 meses","6 a 11 meses","> de 1 año","1 años","2 años","3 años","4 años","5 años"]
            }            
        ],
        'yAxis':[
            {
                'type':"value",
                'axisLine': {
                        'lineStyle': {
                            'color': colors[0]
                        }
                },
            }            
        ],
        'series':[
            {
                'name': 'Cantidad',
                'data': serie,
                'type': "bar",
            }, 
        ], 
    }
    
    return JsonResponse(chart_padron_edades)

#---  PORCENTAJE DE NIÑO SEXO ------------------------------------------------
def get_chart_padron_sexo(_request):
    # Consulta para obtener los datos de ventas
    
    # Formatear los datos para pasarlos a la plantilla
    colors = ['#5470C6', '#91CC75', '#EE6666'];   
    
    serie1 = ['4000']
    serie2 = ['3001']
   
    chart_padron_sexo = {
        'tooltip':{
            'trigger': "item",
        },
        'legend':{
            'top': '5%',
            'left': 'center'# Nombre para la leyenda
        },       
        'series':[
            {
             'name'  : '% de niños sin DNI',
             'type'  : 'pie',
             'radius': ['40%', '70%'],
             'avoidLabelOverlap': 'false',
             'label': {
                'show': 'false',
                'position': 'center'
             },
             'emphasis': {
                 'label': {
                 'show': 'true',
                 'fontSize': 40,
                 'fontWeight': 'bold'
               }
             },
             'labelLine': {
               'show': 'false'
             },
             'data': [
                        { 
                         'value': serie1,
                         'name' : 'Masculino'
                        },
                        { 
                         'value': serie2,
                         'name': 'Femenino' 
                        },
                    ]
            }     
        ] 
    };
    
    return JsonResponse(chart_padron_sexo)

#---  PORCENTAJE DE NIÑO DNI ------------------------------------------------
def get_chart_padron_dni(_request):
    # Consulta para obtener los datos de ventas   
    # Formatear los datos para pasarlos a la plantilla
    colors = ['#5470C6', '#91CC75', '#EE6666'];   
    
    serie1 = ['2700']
    serie2 = ['130789']
   
    chart_padron_dni = {
        'tooltip':{
            'trigger': "item",
        },
        'legend':{
            'top': '2%',
            'left': 'center'# Nombre para la leyenda
        },       
        'series':[
            {
             'name'  : 'Cant niños ',
             'type'  : 'pie',
             'radius': ['40%', '70%'],
             'avoidLabelOverlap': 'false',
             'itemStyle': {
                    'borderRadius': 10,
                    'borderColor': '#fff',
                    'borderWidth': 2
            },
             'label': {
                'show': 'false',
                'position': 'center'
             },
             'emphasis': {
                 'label': {
                 'show': 'true',
                 'fontSize': 40,
                 'fontWeight': 'bold'
               }
             },
             'labelLine': {
               'show': 'false'
             },
             'data': [
                        { 
                         'value': serie1,
                         'name' : 'Sin DNI'
                        },
                        { 
                         'value': serie2,
                         'name': 'Con DNI' 
                        },
                    ]
            }     
        ] 
    };
    
    return JsonResponse(chart_padron_dni)

#---  PORCENTAJE DE NIÑO SIN SEGURO ------------------------------------------------
def get_chart_padron_seguro(_request):
    # Consulta para obtener los datos de ventas   
    # Formatear los datos para pasarlos a la plantilla
    colors = ['#5470C6', '#91CC75', '#EE6666'];   
    
    serie1 = ['4338']
    serie2 = ['129151']
   
    chart_padron_seguro = {
        'tooltip':{
            'trigger': "item",
        },
        'legend':{
            'top': '2%',
            'left': 'center'# Nombre para la leyenda
        },       
        'series':[
            {
             'name'  : 'Cant niños ',
             'type'  : 'pie',
             'radius': ['40%', '70%'],
             'avoidLabelOverlap': 'false',
             'itemStyle': {
                    'borderRadius': 10,
                    'borderColor': '#fff',
                    'borderWidth': 2
            },
             'label': {
                'show': 'false',
                'position': 'center'
             },
             'emphasis': {
                 'label': {
                 'show': 'true',
                 'fontSize': 40,
                 'fontWeight': 'bold'
               }
             },
             'labelLine': {
               'show': 'false'
             },
             'data': [
                        { 
                         'value': serie1,
                         'name' : 'Sin Seguro'
                        },
                        { 
                         'value': serie2,
                         'name': 'Con Seguro' 
                        },
                    ]
            }     
        ] 
    };
    
    return JsonResponse(chart_padron_seguro)

#---  PORCENTAJE DE NIÑO ENCONTRADO ------------------------------------------------
def get_chart_padron_encontrado(_request):
    # Consulta para obtener los datos de ventas   
    # Formatear los datos para pasarlos a la plantilla
    colors = ['#5470C6', '#91CC75', '#EE6666'];   
    
    serie1 = ['4338']
    serie2 = ['129151']
   
    chart_padron_encontrado = {
        'tooltip':{
            'trigger': "item",
        },
        'legend':{
            'top': '2%',
            'left': 'center'# Nombre para la leyenda
        },       
        'series':[
            {
             'name'  : 'Cant niños ',
             'type'  : 'pie',
             'radius': ['40%', '70%'],
             'avoidLabelOverlap': 'false',
             'itemStyle': {
                    'borderRadius': 10,
                    'borderColor': '#fff',
                    'borderWidth': 2
            },
             'label': {
                'show': 'false',
                'position': 'center'
             },
             'emphasis': {
                 'label': {
                 'show': 'true',
                 'fontSize': 40,
                 'fontWeight': 'bold'
               }
             },
             'labelLine': {
               'show': 'false'
             },
             'data': [
                        { 
                         'value': serie1,
                         'name' : 'No Encontrado'
                        },
                        { 
                         'value': serie2,
                         'name': 'Encontrado' 
                        },
                    ]
            }     
        ] 
    };
    
    return JsonResponse(chart_padron_encontrado)

#---  PORCENTAJE DE NIÑO VISITADO ------------------------------------------------
def get_chart_padron_visitado(_request):
    # Consulta para obtener los datos de ventas   
    # Formatear los datos para pasarlos a la plantilla
    colors = ['#5470C6', '#91CC75', '#EE6666'];   
    
    serie1 = ['4338']
    serie2 = ['129151']
   
    chart_padron_visitado = {
        'tooltip':{
            'trigger': "item",
        },
        'legend':{
            'top': '2%',
            'left': 'center'# Nombre para la leyenda
        },       
        'series':[
            {
             'name'  : 'Cant niños ',
             'type'  : 'pie',
             'radius': ['40%', '70%'],
             'avoidLabelOverlap': 'false',
             'itemStyle': {
                    'borderRadius': 10,
                    'borderColor': '#fff',
                    'borderWidth': 2
            },
             'label': {
                'show': 'false',
                'position': 'center'
             },
             'emphasis': {
                 'label': {
                 'show': 'true',
                 'fontSize': 40,
                 'fontWeight': 'bold'
               }
             },
             'labelLine': {
               'show': 'false'
             },
             'data': [
                        { 
                         'value': serie1,
                         'name' : 'Sin Visita'
                        },
                        { 
                         'value': serie2,
                         'name': 'Visitado' 
                        },
                    ]
            }     
        ] 
    };
    
    return JsonResponse(chart_padron_visitado)

#---  PORCENTAJE DE CELULAR MAMA  ------------------------------------------------
def get_chart_padron_celular(_request):
    # Consulta para obtener los datos de ventas   
    # Formatear los datos para pasarlos a la plantilla
    colors = ['#5470C6', '#91CC75', '#EE6666'];   
    
    serie1 = ['4338']
    serie2 = ['129151']
   
    chart_padron_celular = {
        'tooltip':{
            'trigger': "item",
        },
        'legend':{
            'top': '2%',
            'left': 'center'# Nombre para la leyenda
        },       
        'series':[
            {
             'name'  : 'Cant niños ',
             'type'  : 'pie',
             'radius': ['40%', '70%'],
             'avoidLabelOverlap': 'false',
             'itemStyle': {
                    'borderRadius': 10,
                    'borderColor': '#fff',
                    'borderWidth': 2
            },
             'label': {
                'show': 'false',
                'position': 'center'
             },
             'emphasis': {
                 'label': {
                 'show': 'true',
                 'fontSize': 40,
                 'fontWeight': 'bold'
               }
             },
             'labelLine': {
               'show': 'false'
             },
             'data': [
                        { 
                         'value': serie1,
                         'name' : 'Sin Celular Madre'
                        },
                        { 
                         'value': serie2,
                         'name': 'Celular Madre' 
                        },
                    ]
            }     
        ] 
    };
    
    return JsonResponse(chart_padron_celular)

#---  PORCENTAJE DE FECUENCIA  ------------------------------------------------
def get_chart_padron_frecuencia(_request):
    # Consulta para obtener los datos de ventas   
    # Formatear los datos para pasarlos a la plantilla
    colors = ['#5470C6', '#91CC75', '#EE6666'];   
    
    serie1 = ['4338']
    serie2 = ['129151']
   
    chart_padron_frecuencia = {
        'tooltip':{
            'trigger': "item",
        },
        'legend':{
            'top': '2%',
            'left': 'center'# Nombre para la leyenda
        },       
        'series':[
            {
             'name'  : 'Cant niños ',
             'type'  : 'pie',
             'radius': ['40%', '70%'],
             'avoidLabelOverlap': 'false',
             'itemStyle': {
                    'borderRadius': 10,
                    'borderColor': '#fff',
                    'borderWidth': 2
            },
             'label': {
                'show': 'false',
                'position': 'center'
             },
             'emphasis': {
                 'label': {
                 'show': 'true',
                 'fontSize': 40,
                 'fontWeight': 'bold'
               }
             },
             'labelLine': {
               'show': 'false'
             },
             'data': [
                        { 
                         'value': serie1,
                         'name' : 'Sin Frecuencia'
                        },
                        { 
                         'value': serie2,
                         'name': 'Frecuencia Atencion' 
                        },
                    ]
            }     
        ] 
    };
    
    return JsonResponse(chart_padron_frecuencia)

#---  PORCENTAJE DE ENTIDAD  ------------------------------------------------
def get_chart_padron_entidad(_request):
    # Consulta para obtener los datos de ventas   
    # Formatear los datos para pasarlos a la plantilla
    colors = ['#5470C6', '#91CC75', '#EE6666'];   
    
    serie1 = ['4338']
    serie2 = ['9151']
    serie3 = ['3151']
   
    chart_padron_entidad = {
        'tooltip':{
            'trigger': "item",
        },
        'legend':{
            'top': '2%',
            'left': 'center'# Nombre para la leyenda
        },       
        'series':[
            {
             'name'  : 'Cant niños ',
             'type'  : 'pie',
             'radius': ['40%', '70%'],
             'avoidLabelOverlap': 'false',
             'itemStyle': {
                    'borderRadius': 10,
                    'borderColor': '#fff',
                    'borderWidth': 2
            },
             'label': {
                'show': 'false',
                'position': 'center'
             },
             'emphasis': {
                 'label': {
                 'show': 'true',
                 'fontSize': 40,
                 'fontWeight': 'bold'
               }
             },
             'labelLine': {
               'show': 'false'
             },
             'data': [
                        { 
                         'value': serie1,
                         'name' : 'Municipio'
                        },
                        { 
                         'value': serie2,
                         'name': 'EESS SALUD' 
                        },
                                                { 
                         'value': serie3,
                         'name': 'RENIEC' 
                        },
                    ]
            }     
        ] 
    };
    
    return JsonResponse(chart_padron_entidad)

#---  PORCENTAJE DE AVANCE HIS MINSA ------------------------------------------------
def get_chart_padron_atencion(_request):
    # Consulta para obtener los datos de ventas   
    # Formatear los datos para pasarlos a la plantilla
    colors = ['#5470C6', '#91CC75', '#EE6666'];   
    
    serie1 = ['4338']
    serie2 = ['129151']
   
    chart_padron_atencion = {
        'tooltip':{
            'trigger': "item",
        },
        'legend':{
            'top': '2%',
            'left': 'center'# Nombre para la leyenda
        },       
        'series':[
            {
             'name'  : 'Cant niños ',
             'type'  : 'pie',
             'radius': ['40%', '70%'],
             'avoidLabelOverlap': 'false',
             'itemStyle': {
                    'borderRadius': 10,
                    'borderColor': '#fff',
                    'borderWidth': 2
            },
             'label': {
                'show': 'false',
                'position': 'center'
             },
             'emphasis': {
                 'label': {
                 'show': 'true',
                 'fontSize': 40,
                 'fontWeight': 'bold'
               }
             },
             'labelLine': {
               'show': 'false'
             },
             'data': [
                        { 
                         'value': serie1,
                         'name' : 'Sin HIS MINSA'
                        },
                        { 
                         'value': serie2,
                         'name': 'HIS MINSA' 
                        },
                    ]
            }     
        ] 
    };
    
    return JsonResponse(chart_padron_atencion)


#--- PROVINCIAS -------------------------------------------------------------
def get_provincias_padron(request,provincias_id):
    provincias = Provincia.objects.all()
    context = {
                'provincias': provincias,
              }
    return render(request, 'provincias.html', context)


#--- DISTRITOS -------------------------------------------------------------
def get_distritos_padron(request, distritos_id):
    provincias = Provincia.objects.all()
    context = {
                'provincias': provincias,
              }
    return render(request, 'distritos.html',context)

def p_distritos_padron(request):
    provincias = request.GET.get('provincia')
    distritos = Distrito.objects.filter(cod_provincia=provincias)
    context= {
            'provincias': provincias,
            'distritos': distritos
             }  
    return render(request, 'partials/p_distritos.html',context)


################################################
# REPORTE DE SEGUIMIENTO
################################################
class RptProvinciaPadron(TemplateView):
    def get(self,request,*args,**kwargs):
        # variables ingresadas
        edad_inicio = request.GET.get('edad_inicio')
        edad_fin = request.GET.get('edad_fin') 
        provincia = request.GET.get('provincia')
        # creacion de la consulta
        query = Padron.objects.filter(edad_mes__range=[edad_inicio, edad_fin]).filter(ubigeo__startswith=provincia).order_by('provincia','distrito','ap_paterno','ap_materno','nom_nino')
        # Convierte mes a numero
        edad_inicio = int(edad_inicio)
        edad_fin = int(edad_fin)
        
        meses = ["0 años", "1 año", "2 años", "3 años", "4 años", "5 años"]
        nombre_mes_inicio = meses[edad_inicio]
        nombre_mes_fin = meses[edad_fin]      
      
                
        # print(query)
        # creacion de archivo
        wb = Workbook() #crea libro de trabajo
        ws = wb.active #Primera hoja

        # crea titulo del reporte
        ws['B1'].alignment = Alignment(horizontal= "center", vertical="center")
        ws['B1'].font = Font(name = 'Arial', size= 14, bold = True)
        ws['B1'] = 'REPORTE DE SEGUIMIENTO DE NIÑOS Y NIÑAS DEL PADRON NOMINAL POR PROVINCIA'
        # cambina celdas
        ws.merge_cells('B1:U1')

        ws['B3'].alignment = Alignment(horizontal= "left", vertical= "center")
        ws['B3'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B3'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['B3'] = 'Edad Inicio'
        
        ws['C3'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C3'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['C3'].font = Font(name = 'Arial', size= 8)
        ws['C3'].value = nombre_mes_inicio
    
        ws['B4'].alignment = Alignment(horizontal= "left", vertical= "center")
        ws['B4'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B4'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['B4'] = 'Edad Fin'
        
        ws['C4'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C4'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['C4'].font = Font(name = 'Arial', size= 8)
        ws['C4'].number_format = "dd-mm-yyyy"
        ws['C4'].value = nombre_mes_fin
        
        # cambia el alto de la columna
        ws.row_dimensions[1].height = 25
        # cambia el ancho de la columna
        ws.column_dimensions['B'].width = 9
        ws.column_dimensions['C'].width = 9
        ws.column_dimensions['D'].width = 9
        ws.column_dimensions['E'].width = 9
        ws.column_dimensions['F'].width = 9
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 7
        ws.column_dimensions['I'].width = 7
        ws.column_dimensions['J'].width = 7
        ws.column_dimensions['K'].width = 8
        ws.column_dimensions['L'].width = 18
        ws.column_dimensions['M'].width = 18
        ws.column_dimensions['N'].width = 18
        ws.column_dimensions['O'].width = 28
        ws.column_dimensions['P'].width = 20
        ws.column_dimensions['Q'].width = 20
        ws.column_dimensions['R'].width = 16
        ws.column_dimensions['S'].width = 16
        ws.column_dimensions['T'].width = 10
        ws.column_dimensions['U'].width = 10
        ws.column_dimensions['V'].width = 10
        ws.column_dimensions['W'].width = 8
        ws.column_dimensions['X'].width = 20
        ws.column_dimensions['Y'].width = 16
        ws.column_dimensions['Z'].width = 5
        ws.column_dimensions['AA'].width = 9
        ws.column_dimensions['AB'].width = 9
        ws.column_dimensions['AC'].width = 10
        ws.column_dimensions['AD'].width = 10
        ws.column_dimensions['AE'].width = 6
        ws.column_dimensions['AF'].width = 8
        ws.column_dimensions['AG'].width = 18
        ws.column_dimensions['AH'].width = 8
        ws.column_dimensions['AI'].width = 18
        ws.column_dimensions['AJ'].width = 8
        ws.column_dimensions['AK'].width = 18
        ws.column_dimensions['AL'].width = 8
        ws.column_dimensions['AM'].width = 18
        ws.column_dimensions['AN'].width = 18
        ws.column_dimensions['AO'].width = 18
        # linea de division
        #ws.freeze_panes = 'AL8'

        # crea cabecera
        ws['B6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['B6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['B6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['B6'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['B6'] = 'DATOS DEL MENOR DEL PADRON NOMINAL'
        ws.merge_cells('B6:AC6')
        # crea cabecera 2
        ws['AD6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AD6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['AD6'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AD6'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AD6'] = 'DATOS HIS MINSA'
        ws.merge_cells('AD6:AK6')
        # crea cabecera 2
        ws['AL6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AL6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['AL6'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AL6'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AL6'] = 'PERSONAL DE SALUD'
        ws.merge_cells('AL6:AO6')
        
        
        ws['B7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['B7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['B7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['B7'] = 'COD PAD'
        
        ws['C7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C7'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['C7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['C7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['C7'] = 'CNV'

        ws['D7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['D7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['D7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['D7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['D7'] = 'CUI'
        ##
       
        ws['E7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['E7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['E7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['E7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['E7'] = 'DNI'

        ws['F7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['F7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['F7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['F7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['F7'] = 'CNV/DNI'
        
        ws['G7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['G7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['G7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['G7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['G7'] = 'FECHA NAC'
        # celda 
        ws['H7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['H7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['H7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['H7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['H7'] = 'EDAD A'
        # celda        
        ws['I7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['I7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['I7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['I7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['I7'] = 'EDAD M'
        # celda 
        ws['J7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['J7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['J7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['J7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['J7'] = 'EDAD D'
        # celda 
        ws['K7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['K7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['K7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['K7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['K7'] = 'SEGURO'
        # celda 
        ws['L7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['L7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['L7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['L7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['L7'] = 'AP PATERNO'
        # celda 
        ws['M7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['M7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['M7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['M7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['M7'] = 'AP MATERNO'
        # celda 
        ws['N7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['N7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['N7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['N7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['N7'] = 'NOMBRES NIÑO'
        
        ws['O7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['O7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['O7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['O7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['O7'] = 'DIRECCION'
        
        ws['P7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['P7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['P7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['P7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['P7'] = 'EJE'
        
        ws['Q7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['Q7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['Q7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['Q7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['Q7'] = 'REFERENCIA'
        
        ws['R7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['R7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['R7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['R7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['R7'] = 'PROVINCIA'
        
        ws['S7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['S7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['S7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['S7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['S7'] = 'DISTRITO'
        
        ws['T7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['T7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['T7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['T7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['T7'] = 'AREA'
        
        ws['U7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['U7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['U7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['U7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['U7'] = 'VISITADO'
        
        ws['V7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['V7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['V7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['V7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['V7'] = 'FEC VISITA'
        
        ws['W7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['W7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['W7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['W7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['W7'] = 'COD EESS PADRON'
        
        ws['X7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['X7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['X7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['X7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['X7'] = 'NOMBRE EESS PADRON'
        
        ws['Y7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['Y7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['Y7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['Y7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['Y7'] = 'FRECUENCIA'
        
        ws['Z7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['Z7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['Z7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['Z7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['Z7'] = 'ENC'
        
        ws['AA7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AA7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AA7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['AA7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AA7'] = 'DNI MADRE'
        
        ws['AB7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AB7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AB7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['AB7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AB7'] = 'NUM CEL'
        
        ws['AC7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AC7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AC7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['AC7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AC7'] = 'ENTIDAD'
        
        ws['AD7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AD7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AD7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AD7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AD7'] = 'HIS AT'
        
        ws['AE7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AE7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AE7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AE7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AE7'] = 'EDAD AT'
        
        ws['AF7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AF7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AF7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AF7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AF7'] = 'COD RED'
        
        ws['AG7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AG7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AG7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AG7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AG7'] = 'NOM RED'
        
        ws['AH7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AH7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AH7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AH7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AH7'] = 'COD MICRO'
        
        ws['AI7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AI7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AI7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AI7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AI7'] = 'NOM MICRORED'
        
        ws['AJ7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AJ7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AJ7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AJ7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AJ7'] = 'COD EESS'
        
        ws['AK7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AK7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AK7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AK7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AK7'] = 'NOMBRE EESS'
        
        ws['AL7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AL7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AL7'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AL7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AL7'] = 'DNI'
        
        ws['AM7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AM7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AM7'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AM7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AM7'] = 'AP PATERNO'
        
        ws['AN7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AN7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AN7'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AN7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AN7'] = 'AP MATERNO'
        
        ws['AO7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AO7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AO7'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AO7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AO7'] = 'NOMBRE'
        # Pintamos los datos del reporte - RED
        cont = 8      
        
        for q in query:   
            
            # Iconos
            CHECK_ICON = "✔"
            X_ICON = "✖️"

            if q.fecha_nac is not None:
                # Convert the string to a datetime object
                cr_date = datetime.strptime(q.fecha_nac, '%Y-%m-%d')
                # Format the datetime object as 'mm/dd/yyyy'
                fecha_nac_formateada = cr_date.strftime('%m/%d/%Y')
            else:
                fecha_nac_formateada = ''

            
            if q.his_atencion is not None:
                his_date = datetime.strptime(q.his_atencion, '%Y-%m-%d')
                his_ate_formateado = his_date.strftime('%m/%d/%Y')
            else:
                his_ate_formateado = ''
            
            # Icono a usar
            ## icono_visita_1 = CHECK_ICON if q.visita1 == 1 else X_ICON
            ## icono_visita_2 = CHECK_ICON if q.visita2 == 1 else X_ICON
            ## icono_visita_3 = CHECK_ICON if q.visita3 == 1 else X_ICON
            ## icono_visita_4 = CHECK_ICON if q.visita4 == 1 else X_ICON
            
            ws.cell(row = cont , column=2).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=2).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=2).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=2).value = q.cod_padron
            
            ws.cell(row = cont , column=3).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=3).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=3).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=3).value = q.cnv
            
            ws.cell(row = cont , column=4).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=4).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=4).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=4).value = q.cui
            
            ws.cell(row = cont , column=5).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=5).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=5).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=5).value = q.dni
            
            ws.cell(row = cont , column=6).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=6).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=6).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=6).value = q.num_doc
            
            ws.cell(row = cont , column=7).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=7).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=7).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=7).value = fecha_nac_formateada
            
            ws.cell(row = cont , column=8).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=8).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=8).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=8).value = q.edad_anio
            
            ws.cell(row = cont , column=9).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=9).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=9).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=9).value = q.edad_mes
            
            ws.cell(row = cont , column=10).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=10).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=10).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=10).value = q.edad_dias
            
            ws.cell(row = cont , column=11).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=11).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=11).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=11).value = q.seguro
            
            ws.cell(row = cont , column=12).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=12).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=12).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=12).value = q.ap_paterno
            
            ws.cell(row = cont , column=13).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=13).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=13).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=13).value = q.ap_materno
            
            ws.cell(row = cont , column=14).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=14).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=14).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=14).value = q.nom_nino
            
            ws.cell(row = cont , column=15).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=15).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=15).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=15).value = q.direccion
            
            ws.cell(row = cont , column=16).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=16).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=16).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=16).value = q.eje
            
            ws.cell(row = cont , column=17).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=17).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=17).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=17).value = q.referencia
            
            ws.cell(row = cont , column=18).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=18).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=18).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=18).value = q.provincia
            
            ws.cell(row = cont , column=19).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=19).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=19).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=19).value = q.distrito
            
            ws.cell(row = cont , column=20).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=20).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=20).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=20).value = q.area
            
            ws.cell(row = cont , column=21).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=21).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=21).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=21).value = q.visitado
            
            ws.cell(row = cont , column=22).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=22).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=22).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=22).value = q.fe_visita
            
            ws.cell(row = cont , column=23).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=23).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=23).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=23).value = q.cod_eess_padron
            
            ws.cell(row = cont , column=24).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=24).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=24).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=24).value = q.nom_eess_padron
            
            ws.cell(row = cont , column=25).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=25).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=25).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=25).value = q.frecuencia
            
            ws.cell(row = cont , column=26).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=26).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=26).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=26).value = q.encontrado
            
            ws.cell(row = cont , column=27).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=27).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=27).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=27).value = q.dni_mama
            
            ws.cell(row = cont , column=28).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=28).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=28).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=28).value = q.num_cel
            
            ws.cell(row = cont , column=29).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=29).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=29).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=29).value = q.pn_reg
            
            ws.cell(row = cont , column=30).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=30).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=30).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=30).value = his_ate_formateado
            
            ws.cell(row = cont , column=31).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=31).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=31).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=31).value = q.edad_mes_actual
            
            ws.cell(row = cont , column=32).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=32).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=32).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=32).value = q.Cod_Red
            
            ws.cell(row = cont , column=33).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=33).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=33).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=33).value = q.Red
            
            ws.cell(row = cont , column=34).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=34).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=34).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=34).value = q.Cod_Microred
            
            ws.cell(row = cont , column=35).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=35).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=35).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=35).value = q.Microred
            
            ws.cell(row = cont , column=36).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=36).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=36).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=36).value = q.Id_Establecimiento
            
            ws.cell(row = cont , column=37).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=37).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=37).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=37).value = q.Nombre_Establecimiento
            
            ws.cell(row = cont , column=38).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=38).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=38).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=38).value = q.Numero_Documento_Personal
            
            ws.cell(row = cont , column=39).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=39).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=39).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=39).value = q.Apellido_Paterno_Personal
            
            ws.cell(row = cont , column=40).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=40).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=40).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=40).value = q.Apellido_Materno_Personal
            
            ws.cell(row = cont , column=41).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=41).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=41).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=41).value = q.Nombres_Personal
            
            
            cont+=1
               
        #Respuesta de Django
        #Establecer el nombre del archivo
        nombre_archivo = "rpt_seg_visita.xlsx"
        #Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


class RptDistritoPadron(TemplateView):
    def get(self,request,*args,**kwargs):
        # variables ingresadas
        edad_inicio = request.GET.get('edad_inicio')
        edad_fin = request.GET.get('edad_fin') 
        distrito = request.GET.get('distritos')
        # creacion de la consulta
        query = Padron.objects.filter(edad_mes__range=[edad_inicio, edad_fin]).filter(ubigeo=distrito).order_by('provincia','distrito','ap_paterno','ap_materno','nom_nino')
        # Convierte mes a numero
        edad_inicio = int(edad_inicio)
        edad_fin = int(edad_fin)
        
        meses = ["0 años", "1 año", "2 años", "3 años", "4 años", "5 años"]
        nombre_mes_inicio = meses[edad_inicio]
        nombre_mes_fin = meses[edad_fin]      
      
                
        # print(query)
        # creacion de archivo
        wb = Workbook() #crea libro de trabajo
        ws = wb.active #Primera hoja

        # crea titulo del reporte
        ws['B1'].alignment = Alignment(horizontal= "center", vertical="center")
        ws['B1'].font = Font(name = 'Arial', size= 14, bold = True)
        ws['B1'] = 'REPORTE DE SEGUIMIENTO DE NIÑOS Y NIÑAS DEL PADRON NOMINAL POR DISTRITOS'
        # cambina celdas
        ws.merge_cells('B1:U1')

        ws['B3'].alignment = Alignment(horizontal= "left", vertical= "center")
        ws['B3'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B3'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['B3'] = 'Edad Inicio'
        
        ws['C3'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C3'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['C3'].font = Font(name = 'Arial', size= 8)
        ws['C3'].value = nombre_mes_inicio
    
        ws['B4'].alignment = Alignment(horizontal= "left", vertical= "center")
        ws['B4'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B4'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['B4'] = 'Edad Fin'
        
        ws['C4'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C4'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['C4'].font = Font(name = 'Arial', size= 8)
        ws['C4'].number_format = "dd-mm-yyyy"
        ws['C4'].value = nombre_mes_fin
        
        # cambia el alto de la columna
        ws.row_dimensions[1].height = 25
        # cambia el ancho de la columna
        ws.column_dimensions['B'].width = 9
        ws.column_dimensions['C'].width = 9
        ws.column_dimensions['D'].width = 9
        ws.column_dimensions['E'].width = 9
        ws.column_dimensions['F'].width = 9
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 7
        ws.column_dimensions['I'].width = 7
        ws.column_dimensions['J'].width = 7
        ws.column_dimensions['K'].width = 8
        ws.column_dimensions['L'].width = 18
        ws.column_dimensions['M'].width = 18
        ws.column_dimensions['N'].width = 18
        ws.column_dimensions['O'].width = 28
        ws.column_dimensions['P'].width = 20
        ws.column_dimensions['Q'].width = 20
        ws.column_dimensions['R'].width = 16
        ws.column_dimensions['S'].width = 16
        ws.column_dimensions['T'].width = 10
        ws.column_dimensions['U'].width = 10
        ws.column_dimensions['V'].width = 10
        ws.column_dimensions['W'].width = 8
        ws.column_dimensions['X'].width = 20
        ws.column_dimensions['Y'].width = 16
        ws.column_dimensions['Z'].width = 5
        ws.column_dimensions['AA'].width = 9
        ws.column_dimensions['AB'].width = 9
        ws.column_dimensions['AC'].width = 10
        ws.column_dimensions['AD'].width = 10
        ws.column_dimensions['AE'].width = 6
        ws.column_dimensions['AF'].width = 8
        ws.column_dimensions['AG'].width = 18
        ws.column_dimensions['AH'].width = 8
        ws.column_dimensions['AI'].width = 18
        ws.column_dimensions['AJ'].width = 8
        ws.column_dimensions['AK'].width = 18
        ws.column_dimensions['AL'].width = 8
        ws.column_dimensions['AM'].width = 18
        ws.column_dimensions['AN'].width = 18
        ws.column_dimensions['AO'].width = 18
        # linea de division
        #ws.freeze_panes = 'AL8'

        # crea cabecera
        ws['B6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['B6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['B6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['B6'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['B6'] = 'DATOS DEL MENOR DEL PADRON NOMINAL'
        ws.merge_cells('B6:AC6')
        # crea cabecera 2
        ws['AD6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AD6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['AD6'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AD6'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AD6'] = 'DATOS HIS MINSA'
        ws.merge_cells('AD6:AK6')
        # crea cabecera 2
        ws['AL6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AL6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['AL6'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AL6'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AL6'] = 'PERSONAL DE SALUD'
        ws.merge_cells('AL6:AO6')
        
        
        ws['B7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['B7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['B7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['B7'] = 'COD PAD'
        
        ws['C7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C7'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['C7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['C7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['C7'] = 'CNV'

        ws['D7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['D7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['D7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['D7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['D7'] = 'CUI'
        ##
       
        ws['E7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['E7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['E7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['E7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['E7'] = 'DNI'

        ws['F7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['F7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['F7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['F7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['F7'] = 'CNV/DNI'
        
        ws['G7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['G7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['G7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['G7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['G7'] = 'FECHA NAC'
        # celda 
        ws['H7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['H7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['H7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['H7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['H7'] = 'EDAD A'
        # celda        
        ws['I7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['I7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['I7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['I7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['I7'] = 'EDAD M'
        # celda 
        ws['J7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['J7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['J7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['J7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['J7'] = 'EDAD D'
        # celda 
        ws['K7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['K7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['K7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['K7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['K7'] = 'SEGURO'
        # celda 
        ws['L7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['L7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['L7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['L7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['L7'] = 'AP PATERNO'
        # celda 
        ws['M7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['M7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['M7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['M7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['M7'] = 'AP MATERNO'
        # celda 
        ws['N7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['N7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['N7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['N7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['N7'] = 'NOMBRES NIÑO'
        
        ws['O7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['O7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['O7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['O7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['O7'] = 'DIRECCION'
        
        ws['P7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['P7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['P7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['P7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['P7'] = 'EJE'
        
        ws['Q7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['Q7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['Q7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['Q7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['Q7'] = 'REFERENCIA'
        
        ws['R7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['R7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['R7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['R7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['R7'] = 'PROVINCIA'
        
        ws['S7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['S7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['S7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['S7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['S7'] = 'DISTRITO'
        
        ws['T7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['T7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['T7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['T7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['T7'] = 'AREA'
        
        ws['U7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['U7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['U7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['U7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['U7'] = 'VISITADO'
        
        ws['V7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['V7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['V7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['V7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['V7'] = 'FEC VISITA'
        
        ws['W7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['W7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['W7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['W7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['W7'] = 'COD EESS PADRON'
        
        ws['X7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['X7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['X7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['X7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['X7'] = 'NOMBRE EESS PADRON'
        
        ws['Y7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['Y7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['Y7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['Y7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['Y7'] = 'FRECUENCIA'
        
        ws['Z7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['Z7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['Z7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['Z7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['Z7'] = 'ENC'
        
        ws['AA7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AA7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AA7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['AA7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AA7'] = 'DNI MADRE'
        
        ws['AB7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AB7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AB7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['AB7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AB7'] = 'NUM CEL'
        
        ws['AC7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AC7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AC7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['AC7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AC7'] = 'ENTIDAD'
        
        ws['AD7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AD7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AD7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AD7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AD7'] = 'HIS AT'
        
        ws['AE7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AE7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AE7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AE7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AE7'] = 'EDAD AT'
        
        ws['AF7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AF7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AF7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AF7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AF7'] = 'COD RED'
        
        ws['AG7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AG7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AG7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AG7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AG7'] = 'NOM RED'
        
        ws['AH7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AH7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AH7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AH7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AH7'] = 'COD MICRO'
        
        ws['AI7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AI7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AI7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AI7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AI7'] = 'NOM MICRORED'
        
        ws['AJ7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AJ7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AJ7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AJ7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AJ7'] = 'COD EESS'
        
        ws['AK7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AK7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AK7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AK7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AK7'] = 'NOMBRE EESS'
        
        ws['AL7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AL7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AL7'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AL7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AL7'] = 'DNI'
        
        ws['AM7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AM7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AM7'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AM7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AM7'] = 'AP PATERNO'
        
        ws['AN7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AN7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AN7'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AN7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AN7'] = 'AP MATERNO'
        
        ws['AO7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AO7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AO7'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AO7'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['AO7'] = 'NOMBRE'
        # Pintamos los datos del reporte - RED
        cont = 8      
        
        for q in query:   
            
            # Iconos
            CHECK_ICON = "✔"
            X_ICON = "✖️"

            if q.fecha_nac is not None:
                # Convert the string to a datetime object
                cr_date = datetime.strptime(q.fecha_nac, '%Y-%m-%d')
                # Format the datetime object as 'mm/dd/yyyy'
                fecha_nac_formateada = cr_date.strftime('%m/%d/%Y')
            else:
                fecha_nac_formateada = ''

            
            if q.his_atencion is not None:
                his_date = datetime.strptime(q.his_atencion, '%Y-%m-%d')
                his_ate_formateado = his_date.strftime('%m/%d/%Y')
            else:
                his_ate_formateado = ''
            
            # Icono a usar
            ## icono_visita_1 = CHECK_ICON if q.visita1 == 1 else X_ICON
            ## icono_visita_2 = CHECK_ICON if q.visita2 == 1 else X_ICON
            ## icono_visita_3 = CHECK_ICON if q.visita3 == 1 else X_ICON
            ## icono_visita_4 = CHECK_ICON if q.visita4 == 1 else X_ICON
            
            ws.cell(row = cont , column=2).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=2).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=2).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=2).value = q.cod_padron
            
            ws.cell(row = cont , column=3).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=3).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=3).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=3).value = q.cnv
            
            ws.cell(row = cont , column=4).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=4).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=4).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=4).value = q.cui
            
            ws.cell(row = cont , column=5).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=5).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=5).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=5).value = q.dni
            
            ws.cell(row = cont , column=6).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=6).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=6).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=6).value = q.num_doc
            
            ws.cell(row = cont , column=7).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=7).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=7).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=7).value = fecha_nac_formateada
            
            ws.cell(row = cont , column=8).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=8).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=8).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=8).value = q.edad_anio
            
            ws.cell(row = cont , column=9).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=9).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=9).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=9).value = q.edad_mes
            
            ws.cell(row = cont , column=10).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=10).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=10).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=10).value = q.edad_dias
            
            ws.cell(row = cont , column=11).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=11).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=11).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=11).value = q.seguro
            
            ws.cell(row = cont , column=12).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=12).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=12).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=12).value = q.ap_paterno
            
            ws.cell(row = cont , column=13).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=13).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=13).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=13).value = q.ap_materno
            
            ws.cell(row = cont , column=14).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=14).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=14).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=14).value = q.nom_nino
            
            ws.cell(row = cont , column=15).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=15).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=15).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=15).value = q.direccion
            
            ws.cell(row = cont , column=16).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=16).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=16).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=16).value = q.eje
            
            ws.cell(row = cont , column=17).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=17).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=17).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=17).value = q.referencia
            
            ws.cell(row = cont , column=18).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=18).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=18).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=18).value = q.provincia
            
            ws.cell(row = cont , column=19).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=19).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=19).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=19).value = q.distrito
            
            ws.cell(row = cont , column=20).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=20).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=20).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=20).value = q.area
            
            ws.cell(row = cont , column=21).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=21).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=21).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=21).value = q.visitado
            
            ws.cell(row = cont , column=22).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=22).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=22).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=22).value = q.fe_visita
            
            ws.cell(row = cont , column=23).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=23).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=23).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=23).value = q.cod_eess_padron
            
            ws.cell(row = cont , column=24).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=24).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=24).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=24).value = q.nom_eess_padron
            
            ws.cell(row = cont , column=25).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=25).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=25).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=25).value = q.frecuencia
            
            ws.cell(row = cont , column=26).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=26).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=26).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=26).value = q.encontrado
            
            ws.cell(row = cont , column=27).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=27).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=27).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=27).value = q.dni_mama
            
            ws.cell(row = cont , column=28).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=28).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=28).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=28).value = q.num_cel
            
            ws.cell(row = cont , column=29).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=29).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=29).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=29).value = q.pn_reg
            
            ws.cell(row = cont , column=30).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=30).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=30).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=30).value = his_ate_formateado
            
            ws.cell(row = cont , column=31).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=31).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=31).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=31).value = q.edad_mes_actual
            
            ws.cell(row = cont , column=32).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=32).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=32).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=32).value = q.Cod_Red
            
            ws.cell(row = cont , column=33).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=33).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=33).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=33).value = q.Red
            
            ws.cell(row = cont , column=34).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=34).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=34).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=34).value = q.Cod_Microred
            
            ws.cell(row = cont , column=35).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=35).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=35).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=35).value = q.Microred
            
            ws.cell(row = cont , column=36).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=36).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=36).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=36).value = q.Id_Establecimiento
            
            ws.cell(row = cont , column=37).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=37).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=37).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=37).value = q.Nombre_Establecimiento
            
            ws.cell(row = cont , column=38).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=38).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=38).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=38).value = q.Numero_Documento_Personal
            
            ws.cell(row = cont , column=39).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=39).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=39).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=39).value = q.Apellido_Paterno_Personal
            
            ws.cell(row = cont , column=40).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=40).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=40).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=40).value = q.Apellido_Materno_Personal
            
            ws.cell(row = cont , column=41).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=41).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=41).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=41).value = q.Nombres_Personal
            
            
            cont+=1
               
        #Respuesta de Django
        #Establecer el nombre del archivo
        nombre_archivo = "rpt_seg_visita.xlsx"
        #Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


################################################
# ACTAS PADRON NOMINAL 
################################################
@login_required
def acta_index(request):
    p_chyo = 'CHANCHAMAYO'
    p_chupaca = 'CHUPACA'
    p_concepcion = 'CONCEPCION'
    p_huancayo = 'HUANCAYO'
    p_jauja = 'JAUJA'
    p_junin = 'JUNIN'
    p_satipo = 'SATIPO'
    p_tarma = 'TARMA'
    p_yauli = 'YAULI'
       
    t_chyo = Padron.objects.values('distrito').annotate( 
                                                        sum_0_28d =Sum('edad_0_28_dias'),
                                                        sum_0_5m =Sum('edad_0_5_meses'),
                                                        sum_6_11m=Sum('edad_6_11_meses'),
                                                        sum_12m=Sum('menores_de_12_meses'),
                                                        sum_1a=Sum('edad_1_anio'),
                                                        sum_2a=Sum('edad_2_anios'),
                                                        sum_3a=Sum('edad_3_anios'),
                                                        sum_4a=Sum('edad_4_anios'),
                                                        sum_5a=Sum('edad_5_anios'),
                                                        sum_num = Count('num_num_doc', filter=Q(num_num_doc=0)),
                                                        sum_seguro= Count('num_seguro',filter=Q(num_seguro=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0)),
                                                        sum_frecuencia= Count('num_frecuencia', filter=Q(num_frecuencia=0)),
                                                        sum_entidad= Count('num_entidad_reniec', filter=Q(num_entidad_reniec=1)),
                                                        ).filter(provincia=p_chyo).order_by('distrito')                                                                                   

    t_chupaca = Padron.objects.values('distrito').annotate( 
                                                        sum_0_28d =Sum('edad_0_28_dias'),
                                                        sum_0_5m =Sum('edad_0_5_meses'),
                                                        sum_6_11m=Sum('edad_6_11_meses'),
                                                        sum_12m=Sum('menores_de_12_meses'),
                                                        sum_1a=Sum('edad_1_anio'),
                                                        sum_2a=Sum('edad_2_anios'),
                                                        sum_3a=Sum('edad_3_anios'),
                                                        sum_4a=Sum('edad_4_anios'),
                                                        sum_5a=Sum('edad_5_anios'),
                                                        sum_num = Count('num_num_doc', filter=Q(num_num_doc=0)),
                                                        sum_seguro= Count('num_seguro',filter=Q(num_seguro=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0)),
                                                        sum_frecuencia= Count('num_frecuencia', filter=Q(num_frecuencia=0)),
                                                        sum_entidad= Count('num_entidad_reniec', filter=Q(num_entidad_reniec=1)),
                                                        ).filter(provincia=p_chupaca).order_by('distrito')                                          
    
    t_concepcion = Padron.objects.values('distrito').annotate( 
                                                        sum_0_28d =Sum('edad_0_28_dias'),
                                                        sum_0_5m =Sum('edad_0_5_meses'),
                                                        sum_6_11m=Sum('edad_6_11_meses'),
                                                        sum_12m=Sum('menores_de_12_meses'),
                                                        sum_1a=Sum('edad_1_anio'),
                                                        sum_2a=Sum('edad_2_anios'),
                                                        sum_3a=Sum('edad_3_anios'),
                                                        sum_4a=Sum('edad_4_anios'),
                                                        sum_5a=Sum('edad_5_anios'),
                                                        sum_num = Count('num_num_doc', filter=Q(num_num_doc=0)),
                                                        sum_seguro= Count('num_seguro',filter=Q(num_seguro=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0)),
                                                        sum_frecuencia= Count('num_frecuencia', filter=Q(num_frecuencia=0)),
                                                        sum_entidad= Count('num_entidad_reniec', filter=Q(num_entidad_reniec=1)),
                                                        ).filter(provincia=p_concepcion).order_by('distrito')                                                        
    
    t_huancayo = Padron.objects.values('distrito').annotate( 
                                                        sum_0_28d =Sum('edad_0_28_dias'),
                                                        sum_0_5m =Sum('edad_0_5_meses'),
                                                        sum_6_11m=Sum('edad_6_11_meses'),
                                                        sum_12m=Sum('menores_de_12_meses'),
                                                        sum_1a=Sum('edad_1_anio'),
                                                        sum_2a=Sum('edad_2_anios'),
                                                        sum_3a=Sum('edad_3_anios'),
                                                        sum_4a=Sum('edad_4_anios'),
                                                        sum_5a=Sum('edad_5_anios'),
                                                        sum_num = Count('num_num_doc', filter=Q(num_num_doc=0)),
                                                        sum_seguro= Count('num_seguro',filter=Q(num_seguro=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0)),
                                                        sum_frecuencia= Count('num_frecuencia', filter=Q(num_frecuencia=0)),
                                                        sum_entidad= Count('num_entidad_reniec', filter=Q(num_entidad_reniec=1)),
                                                        ).filter(provincia=p_huancayo).order_by('distrito')      
    
    t_jauja = Padron.objects.values('distrito').annotate( 
                                                        sum_0_28d =Sum('edad_0_28_dias'),
                                                        sum_0_5m =Sum('edad_0_5_meses'),
                                                        sum_6_11m=Sum('edad_6_11_meses'),
                                                        sum_12m=Sum('menores_de_12_meses'),
                                                        sum_1a=Sum('edad_1_anio'),
                                                        sum_2a=Sum('edad_2_anios'),
                                                        sum_3a=Sum('edad_3_anios'),
                                                        sum_4a=Sum('edad_4_anios'),
                                                        sum_5a=Sum('edad_5_anios'),
                                                        sum_num = Count('num_num_doc', filter=Q(num_num_doc=0)),
                                                        sum_seguro= Count('num_seguro',filter=Q(num_seguro=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0)),
                                                        sum_frecuencia= Count('num_frecuencia', filter=Q(num_frecuencia=0)),
                                                        sum_entidad= Count('num_entidad_reniec', filter=Q(num_entidad_reniec=1)),
                                                        ).filter(provincia=p_jauja).order_by('distrito')      
    
    t_junin = Padron.objects.values('distrito').annotate( 
                                                        sum_0_28d =Sum('edad_0_28_dias'),
                                                        sum_0_5m =Sum('edad_0_5_meses'),
                                                        sum_6_11m=Sum('edad_6_11_meses'),
                                                        sum_12m=Sum('menores_de_12_meses'),
                                                        sum_1a=Sum('edad_1_anio'),
                                                        sum_2a=Sum('edad_2_anios'),
                                                        sum_3a=Sum('edad_3_anios'),
                                                        sum_4a=Sum('edad_4_anios'),
                                                        sum_5a=Sum('edad_5_anios'),
                                                        sum_num = Count('num_num_doc', filter=Q(num_num_doc=0)),
                                                        sum_seguro= Count('num_seguro',filter=Q(num_seguro=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0)),
                                                        sum_frecuencia= Count('num_frecuencia', filter=Q(num_frecuencia=0)),
                                                        sum_entidad= Count('num_entidad_reniec', filter=Q(num_entidad_reniec=1)),
                                                        ).filter(provincia=p_junin).order_by('distrito') 
    
    t_satipo = Padron.objects.values('distrito').annotate( 
                                                        sum_0_28d =Sum('edad_0_28_dias'),
                                                        sum_0_5m =Sum('edad_0_5_meses'),
                                                        sum_6_11m=Sum('edad_6_11_meses'),
                                                        sum_12m=Sum('menores_de_12_meses'),
                                                        sum_1a=Sum('edad_1_anio'),
                                                        sum_2a=Sum('edad_2_anios'),
                                                        sum_3a=Sum('edad_3_anios'),
                                                        sum_4a=Sum('edad_4_anios'),
                                                        sum_5a=Sum('edad_5_anios'),
                                                        sum_num = Count('num_num_doc', filter=Q(num_num_doc=0)),
                                                        sum_seguro= Count('num_seguro',filter=Q(num_seguro=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0)),
                                                        sum_frecuencia= Count('num_frecuencia', filter=Q(num_frecuencia=0)),
                                                        sum_entidad= Count('num_entidad_reniec', filter=Q(num_entidad_reniec=1)),
                                                        ).filter(provincia=p_satipo).order_by('distrito')        
    
    t_tarma = Padron.objects.values('distrito').annotate( 
                                                        sum_0_28d =Sum('edad_0_28_dias'),
                                                        sum_0_5m =Sum('edad_0_5_meses'),
                                                        sum_6_11m=Sum('edad_6_11_meses'),
                                                        sum_12m=Sum('menores_de_12_meses'),
                                                        sum_1a=Sum('edad_1_anio'),
                                                        sum_2a=Sum('edad_2_anios'),
                                                        sum_3a=Sum('edad_3_anios'),
                                                        sum_4a=Sum('edad_4_anios'),
                                                        sum_5a=Sum('edad_5_anios'),
                                                        sum_num = Count('num_num_doc', filter=Q(num_num_doc=0)),
                                                        sum_seguro= Count('num_seguro',filter=Q(num_seguro=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0)),
                                                        sum_frecuencia= Count('num_frecuencia', filter=Q(num_frecuencia=0)),
                                                        sum_entidad= Count('num_entidad_reniec', filter=Q(num_entidad_reniec=1)),
                                                        ).filter(provincia=p_tarma).order_by('distrito')  
    
    t_yauli = Padron.objects.values('distrito').annotate( 
                                                        sum_0_28d =Sum('edad_0_28_dias'),
                                                        sum_0_5m =Sum('edad_0_5_meses'),
                                                        sum_6_11m=Sum('edad_6_11_meses'),
                                                        sum_12m=Sum('menores_de_12_meses'),
                                                        sum_1a=Sum('edad_1_anio'),
                                                        sum_2a=Sum('edad_2_anios'),
                                                        sum_3a=Sum('edad_3_anios'),
                                                        sum_4a=Sum('edad_4_anios'),
                                                        sum_5a=Sum('edad_5_anios'),
                                                        sum_num = Count('num_num_doc', filter=Q(num_num_doc=0)),
                                                        sum_seguro= Count('num_seguro',filter=Q(num_seguro=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0)),
                                                        sum_frecuencia= Count('num_frecuencia', filter=Q(num_frecuencia=0)),
                                                        sum_entidad= Count('num_entidad_reniec', filter=Q(num_entidad_reniec=1)),
                                                        ).filter(provincia=p_yauli).order_by('distrito')  
    
    context = {
                't_chyo'      : t_chyo,
                't_chupaca'   : t_chupaca,
                't_concepcion': t_concepcion,
                't_huancayo'  : t_huancayo,
                't_jauja'     : t_jauja,
                't_junin'     : t_junin,
                't_satipo'    : t_satipo,
                't_tarma'     : t_tarma,
                't_yauli'     : t_yauli,
              }
    
    return render(request, 'actas.html', context)


################################################
# ACTAS PADRON NOMINAL 
################################################
@login_required
def compromiso_index(request):
    p_chyo = 'CHANCHAMAYO'
    p_chupaca = 'CHUPACA'
    p_concepcion = 'CONCEPCION'
    p_huancayo = 'HUANCAYO'
    p_jauja = 'JAUJA'
    p_junin = 'JUNIN'
    p_satipo = 'SATIPO'
    p_tarma = 'TARMA'
    p_yauli = 'YAULI'
       
    t_chyo = Compromiso.objects.values('distrito').annotate( 
                                                        sum_num =Sum('num'),
                                                        sum_den =Sum('den'),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0))
                                                        ).annotate(
                                                            porcentaje=ExpressionWrapper(
                                                                    (F('sum_num') * 100.0) / F('sum_den'),
                                                                    output_field=FloatField()
                                                                )                                                      
                                                        ).filter(provincia=p_chyo).order_by('distrito')                                                                                   

    t_chupaca = Compromiso.objects.values('distrito').annotate( 
                                                        sum_num =Sum('num'),
                                                        sum_den =Sum('den'),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0))
                                                        ).annotate(
                                                            porcentaje=ExpressionWrapper(
                                                                    (F('sum_num') * 100.0) / F('sum_den'),
                                                                    output_field=FloatField()
                                                                )                                                      
                                                        ).filter(provincia=p_chupaca).order_by('distrito')                                          
    
    t_concepcion = Compromiso.objects.values('distrito').annotate( 
                                                        sum_num =Sum('num'),
                                                        sum_den =Sum('den'),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0))
                                                        ).annotate(
                                                            porcentaje=ExpressionWrapper(
                                                                    (F('sum_num') * 100.0) / F('sum_den'),
                                                                    output_field=FloatField()
                                                                )                                                      
                                                        ).filter(provincia=p_concepcion).order_by('distrito')                                                      
    
    t_huancayo = Compromiso.objects.values('distrito').annotate( 
                                                        sum_num =Sum('num'),
                                                        sum_den =Sum('den'),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0))
                                                        ).annotate(
                                                            porcentaje=ExpressionWrapper(
                                                                    (F('sum_num') * 100.0) / F('sum_den'),
                                                                    output_field=FloatField()
                                                                )                                                      
                                                        ).filter(provincia=p_huancayo).order_by('distrito')    
    
    t_jauja = Compromiso.objects.values('distrito').annotate( 
                                                        sum_num =Sum('num'),
                                                        sum_den =Sum('den'),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0))
                                                        ).annotate(
                                                            porcentaje=ExpressionWrapper(
                                                                    (F('sum_num') * 100.0) / F('sum_den'),
                                                                    output_field=FloatField()
                                                                )                                                      
                                                        ).filter(provincia=p_jauja).order_by('distrito')    
    
    t_junin = Compromiso.objects.values('distrito').annotate( 
                                                        sum_num =Sum('num'),
                                                        sum_den =Sum('den'),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0))
                                                        ).annotate(
                                                            porcentaje=ExpressionWrapper(
                                                                    (F('sum_num') * 100.0) / F('sum_den'),
                                                                    output_field=FloatField()
                                                                )                                                      
                                                        ).filter(provincia=p_junin).order_by('distrito')  
    
    t_satipo = Compromiso.objects.values('distrito').annotate( 
                                                        sum_num =Sum('num'),
                                                        sum_den =Sum('den'),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0))
                                                        ).annotate(
                                                            porcentaje=ExpressionWrapper(
                                                                    (F('sum_num') * 100.0) / F('sum_den'),
                                                                    output_field=FloatField()
                                                                )                                                      
                                                        ).filter(provincia=p_satipo).order_by('distrito')        
    
    t_tarma = Compromiso.objects.values('distrito').annotate( 
                                                        sum_num =Sum('num'),
                                                        sum_den =Sum('den'),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0))
                                                        ).annotate(
                                                            porcentaje=ExpressionWrapper(
                                                                    (F('sum_num') * 100.0) / F('sum_den'),
                                                                    output_field=FloatField()
                                                                )                                                      
                                                        ).filter(provincia=p_tarma).order_by('distrito')  
    
    t_yauli = Compromiso.objects.values('distrito').annotate( 
                                                        sum_num =Sum('num'),
                                                        sum_den =Sum('den'),
                                                        sum_vis= Count('num_vis', filter=Q(num_vis=0)),
                                                        sum_enc= Count('num_enc', filter=Q(num_enc=0)),
                                                        sum_num_cel= Count('num_num_cel', filter=Q(num_num_cel=0))
                                                        ).annotate(
                                                            porcentaje=ExpressionWrapper(
                                                                    (F('sum_num') * 100.0) / F('sum_den'),
                                                                    output_field=FloatField()
                                                                )                                                      
                                                        ).filter(provincia=p_yauli).order_by('distrito')  
    
    context = {
                't_chyo'      : t_chyo,
                't_chupaca'   : t_chupaca,
                't_concepcion': t_concepcion,
                't_huancayo'  : t_huancayo,
                't_jauja'     : t_jauja,
                't_junin'     : t_junin,
                't_satipo'    : t_satipo,
                't_tarma'     : t_tarma,
                't_yauli'     : t_yauli,
              }
    
    return render(request, 'compromiso.html', context)
