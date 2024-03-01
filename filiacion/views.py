from django.shortcuts import render, redirect, get_object_or_404
from django.http.response import JsonResponse
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.db import IntegrityError
from django.contrib.auth.decorators import login_required

from django.core.files.storage import FileSystemStorage

# Create your views here.
from .models import Filiacion, Directorio, DirectorioRed, DirectorioEstablecimiento, Diresa, Provincia, Distrito, Red, Microred, Establecimiento 
from .models import rpt_certificado, ActualizaBD, RptVisitaDis, RptSeguimientoVisitaDis,RptVisita, TipoReporte, Item_mes

# report excel
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# report operacionales
from django.db import models
from django.db import connection
from django.views import View

# graficos
from random import randrange

# tablas por redes
from django.db.models import Sum, F, FloatField, ExpressionWrapper
from django.db.models.functions import Cast, Round

# ht-get por distrito
from django.urls import reverse
import json

def home(request):
    actualiza = ActualizaBD.objects.all()
    context = {
                'actualiza': actualiza,
                }    
    return render(request, 'home.html', context)

@login_required
def home_dashboard(request):
    actualiza = ActualizaBD.objects.all()
    context = {
                'actualiza': actualiza,
                }    
    return render(request, 'home_dashboard.html', context)

# ----- DIRECTORIO MUNICIPIO --------------------
@login_required
def filiacion(request):
    filiaciones = Filiacion.objects.all()
    context = {
                'filiaciones': filiaciones,
                }
    return render(request, 'filiacion.html', context)

# ----- DIRECTORIO SALUD RED --------------------
@login_required
def directorio_red(request):
    directorio_redes = DirectorioRed.objects.all()
    context = {
                'directorio_redes': directorio_redes,
                }
    return render(request, 'directorio_red.html', context)

# ----- DIRECTORIO SALUD ESTABLECIMIENTO --------------------
@login_required
def directorio_establecimiento(request):
    directorio_establecimientos = DirectorioEstablecimiento.objects.all()
    context = {
                'directorio_establecimientos': directorio_establecimientos,
                }
    return render(request, 'directorio_establecimiento.html', context)

# ----- INICIO DE SESION --------------------------------
@login_required
def signout(request):
    logout(request)
    return redirect('home')

def signin(request):
    if request.method == 'GET':
        return render(request, 'signin.html', {"form": AuthenticationForm})
    else:
        user = authenticate(
            request, username=request.POST['username'], password=request.POST['password'])
        if user is None:
            return render(request, 'signin.html', {"form": AuthenticationForm, "error": "Username or password is incorrect."})

        login(request, user)
        return redirect('home_dashboard')

def signup(request):
    if request.method == 'GET':
        return render(request, 'signup.html', {
            'form': UserCreationForm
        })
    else:
        if request.POST['password1'] == request.POST['password2']:
            try:
                user = User.objects.create_user(
                    username=request.POST['username'], password=request.POST['password1'])
                user.save()
                login(request, user)
                return redirect('home_dashboard')
            except IntegrityError:
                return render(request, 'signup.html', {
                    'form': UserCreationForm,
                    "error": 'Usuario ya existe'
                })
        return render(request, 'signup.html', {
            'form': UserCreationForm,
            "error": 'Password fo not match'
        })

# ----- SELECT DEPENDIENTES FILIACION --------------------
def distrito(request):
    provincias = request.GET.get('provincia_selected')
    distritos = Distrito.objects.filter(provincia_id=provincias)
    context = {
                'distritos': distritos, 
                'is_htmx': True 
                }
    return render(request, 'partials/distritos.html', context)

# ----- FRONTEND FILIACION --------------------
def frontend_filiacion(request):
    filiaciones = Filiacion.objects.all()
    context = {
            'filiaciones': filiaciones,
            }
    return render(request, 'frontend/filiacion.html', context)

def frontend_directorio_diresa(request):
    directorio_diresas= Directorio.objects.all()
    context = {
            'directorio_diresas': directorio_diresas,
            }
    return render(request, 'frontend/directorio_diresa.html', context)

def frontend_directorio_red(request):
    directorio_redes= DirectorioRed.objects.all()
    context = {
            'directorio_redes': directorio_redes,
            }
    return render(request, 'frontend/directorio_red.html', context)

def frontend_directorio_establecimiento(request):
    directorio_establecimientos= DirectorioEstablecimiento.objects.all()
    context = {
            'directorio_establecimientos': directorio_establecimientos,
            }
    return render(request, 'frontend/directorio_establecimiento.html', context)

#############################################
# ----- RPT DISCAPACIDAD --------------------
#############################################
@login_required
def listar_rpt_discapacidad(request):
    
    # Obtener el filtro de mes y año del parámetro GET
    # Obtener todas las marcaciones o filtrar por mes/año
    return render(request, 'rpt_discapacidad/rpt_discapacidad.html')

class ReportePersonalizadoExcel(TemplateView):
    def get(self,request,*args,**kwargs):
        # creacion de la consulta
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        query = rpt_certificado.objects.filter(Fecha_Atencion__range=[fecha_inicio, fecha_fin]).order_by('Red','MicroRed','Nombre_Establecimiento')

        # creacion de archivo
        wb = Workbook() #crea libro de trabajo
        ws = wb.active #Primera hoja

        # crea titulo del reporte
        ws['A1'].alignment = Alignment(horizontal= "center", vertical="center")
        ws['A1'].font = Font(name = 'Arial', size= 14, bold = True)
        ws['A1'] = 'REPORTE DE CERTIFICADOS DE DISCAPACIDAD'
        # cambina celdas
        ws.merge_cells('A1:K1')

        ws['B3'].alignment = Alignment(horizontal= "left", vertical= "center")
        ws['B3'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B3'].font = Font(name = 'Arial', size= 8)
        ws['B3'] = 'Fecha Inicio'
        
        ws['C3'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C3'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['C3'].font = Font(name = 'Arial', size= 8)
        ws['C3'].value = fecha_inicio
    
        ws['B4'].alignment = Alignment(horizontal= "left", vertical= "center")
        ws['B4'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B4'].font = Font(name = 'Arial', size= 8)
        ws['B4'] = 'Fecha Fin'
        
        ws['C4'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C4'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['C4'].font = Font(name = 'Arial', size= 8)
        ws['C4'].number_format = "dd-mm-yyyy"
        ws['C4'].value = fecha_fin
        
        # cambia el alto de la columna
        ws.row_dimensions[1].height = 25
        # cambia el ancho de la columna
        ws.column_dimensions['B'].width = 23
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 32
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 10
        # linea de division
        ws.freeze_panes = 'AL8'

        # crea cabecera
        # celda red 
        ws['B6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['B6'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['B6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['B6'] = 'RED'
        ws.merge_cells('B6:B7')
        # celda microred 
        ws['C6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['C6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['C6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['C6'] = 'MICRORED'
        ws.merge_cells('C6:C7')
        # celda establecimiento
        ws['D6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['D6'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['D6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['D6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['D6'] = 'COD ESTABLEC'
        ws.merge_cells('D6:D7')
        # celda codigo de establecimiento
        ws['E6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['E6'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['E6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['E6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['E6'] = 'NOMBRE ESTABLECIMIENTO'
        ws.merge_cells('E6:E7')

        # celda codigo de establecimiento
        ws['F6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['F6'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['F6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['F6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['F6'] = 'EVALUACION'
        ws.merge_cells('F6:F7')
        # celda codigo de establecimiento
        ws['G6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['G6'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['G6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['G6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['G6'] = 'CALIFICACION'
        ws.merge_cells('G6:G7')
        # celda TITULO
        ws['H6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['H6'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['H6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['H6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['H6'] = 'CERTIFICACION'
        ws.merge_cells('H6:K6')
        # celda 
        ws['H7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['H7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['H7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['H7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['H7'] = 'LEVE'
        # celda 
        ws['I7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['I7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['I7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['I7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['I7'] = 'MODERADO'
        # celda 
        ws['J7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['J7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['J7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['J7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['J7'] = 'SEVERO'
        # celda 
        ws['K7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['K7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['K7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['K7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['K7'] = 'S/GRADO'

        # Pintamos los datos del reporte - RED
        cont = 8       
        for q in query:   
            ws.cell(row = cont , column=2).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=2).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=2).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column=2).value = q.Red

            ws.cell(row = cont , column=3).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=3).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=3).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column=3).value = q.MicroRed
            
            ws.cell(row = cont , column=4).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=4).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=4).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column= 4).value = q.Codigo_Unico
            
            ws.cell(row = cont , column=5).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=5).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=5).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column= 5).value = q.Nombre_Establecimiento
            
            ws.cell(row = cont , column=6).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=6).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=6).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column= 6).value = q.DIS_EVALUACION
            
            ws.cell(row = cont , column=7).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=7).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=7).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column= 7).value = q.DIS_CALIFICACION
            
            ws.cell(row = cont , column=8).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=8).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=8).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column= 8).value = q.DIS_LEV
            
            ws.cell(row = cont , column=9).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=9).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=9).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column= 9).value = q.DIS_MOD
            
            ws.cell(row = cont , column=10).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=10).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=10).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column= 10).value = q.DIS_SEV
            
            ws.cell(row = cont , column=11).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=11).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=11).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column= 11).value = q.DIS_TOTAL            
            cont+=1

                
        #Respuesta de Django
        #Establecer el nombre del archivo
        nombre_archivo = "rpt_discapacidad.xlsx"
        #Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

#############################################
# ----- RPT VISITA  --------------------
#############################################
@login_required
def listar_rpt_visita_dis(request):
    
    # Obtener el filtro de mes y año del parámetro GET
    # Obtener todas las marcaciones o filtrar por mes/año
    return render(request, 'rpt_discapacidad/rpt_visita_dis.html')

class RptVistaDisExcel(TemplateView):
    def get(self,request,*args,**kwargs):
        # creacion de la consulta
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        query = RptVisitaDis.objects.filter(Fecha_Atencion__range=[fecha_inicio, fecha_fin]).order_by('Red','MicroRed','Nombre_Establecimiento')

        # creacion de archivo
        wb = Workbook() #crea libro de trabajo
        ws = wb.active #Primera hoja

        # crea titulo del reporte
        ws['A1'].alignment = Alignment(horizontal= "center", vertical="center")
        ws['A1'].font = Font(name = 'Arial', size= 14, bold = True)
        ws['A1'] = 'REPORTE DE VISITA DOMICILIARIA A PACIENTES CON DISCAPACIDAD'
        # cambina celdas
        ws.merge_cells('A1:J1')

        ws['B3'].alignment = Alignment(horizontal= "left", vertical= "center")
        ws['B3'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B3'].font = Font(name = 'Arial', size= 8)
        ws['B3'] = 'Fecha Inicio'
        
        ws['C3'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C3'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['C3'].font = Font(name = 'Arial', size= 8)
        ws['C3'].value = fecha_inicio
    
        ws['B4'].alignment = Alignment(horizontal= "left", vertical= "center")
        ws['B4'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B4'].font = Font(name = 'Arial', size= 8)
        ws['B4'] = 'Fecha Fin'
        
        ws['C4'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C4'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['C4'].font = Font(name = 'Arial', size= 8)
        ws['C4'].number_format = "dd-mm-yyyy"
        ws['C4'].value = fecha_fin
        
        # cambia el alto de la columna
        ws.row_dimensions[1].height = 25
        # cambia el ancho de la columna
        ws.column_dimensions['B'].width = 23
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 32
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10
        # linea de division
        ws.freeze_panes = 'AL8'

        # crea cabecera
        ws['B6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['B6'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['B6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['B6'] = 'RED'
        ws.merge_cells('B6:B7')

        ws['C6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['C6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['C6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['C6'] = 'MICRORED'
        ws.merge_cells('C6:C7')

        ws['D6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['D6'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['D6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['D6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['D6'] = 'COD ESTABLEC'
        ws.merge_cells('D6:D7')

        ws['E6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['E6'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['E6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['E6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['E6'] = 'NOMBRE ESTABLECIMIENTO'
        ws.merge_cells('E6:E7')

        ws['F6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['F6'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['F6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['F6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['F6'] = '1° VISITA'
        ws.merge_cells('F6:F7')

        ws['G6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['G6'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['G6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['G6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['G6'] = '2° VISITA'
        ws.merge_cells('G6:G7')
        # celda 
        ws['H6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['H6'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['H6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['H6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['H6'] = '3° VISITA'
        ws.merge_cells('H6:H7')
        # celda 
        ws['I6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['I6'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['I6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['I6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['I6'] = '4° VISITA'
        ws.merge_cells('I6:I7')
        # celda 

        # Pintamos los datos del reporte - RED
        cont = 8       
        for q in query:   
            ws.cell(row = cont , column=2).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=2).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=2).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column=2).value = q.Red

            ws.cell(row = cont , column=3).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=3).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=3).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column=3).value = q.MicroRed
            
            ws.cell(row = cont , column=4).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=4).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=4).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column= 4).value = q.Codigo_Unico
            
            ws.cell(row = cont , column=5).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=5).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=5).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column= 5).value = q.Nombre_Establecimiento
            
            ws.cell(row = cont , column=6).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=6).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=6).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column= 6).value = q.VISITA_1
            
            ws.cell(row = cont , column=7).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=7).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=7).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column= 7).value = q.VISITA_2
            
            ws.cell(row = cont , column=8).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=8).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=8).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column= 8).value = q.VISITA_3
            
            ws.cell(row = cont , column=9).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=9).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=9).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column= 9).value = q.VISITA_4
                     
            cont+=1

                
        #Respuesta de Django
        #Establecer el nombre del archivo
        nombre_archivo = "rpt_visita_dis.xlsx"
        #Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

#############################################
# ----- RPT SEGUMIENTO VISITA  --------------
#############################################
@login_required
def listar_rpt_seguimiento_visita_dis(request):
    
    # Obtener el filtro de mes y año del parámetro GET
    # Obtener todas las marcaciones o filtrar por mes/año
    return render(request, 'rpt_discapacidad/rpt_seguimiento_visita_dis.html')

class RptSeguimientoVistaDisExcel(TemplateView):
    def get(self,request,*args,**kwargs):
        # creacion de la consulta
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        query = RptSeguimientoVisitaDis.objects.filter(FECHA_VISITA_1__range=[fecha_inicio, fecha_fin]).order_by('Red','MicroRed','Nombre_Establecimiento')

        # creacion de archivo
        wb = Workbook() #crea libro de trabajo
        ws = wb.active #Primera hoja

        # crea titulo del reporte
        ws['A1'].alignment = Alignment(horizontal= "center", vertical="center")
        ws['A1'].font = Font(name = 'Arial', size= 14, bold = True)
        ws['A1'] = 'REPORTE DE SEGUIMIENTO VISITA DOMICILIARIA A PACIENTES CON DISCAPACIDAD'
        # cambina celdas
        ws.merge_cells('A1:J1')

        ws['B3'].alignment = Alignment(horizontal= "left", vertical= "center")
        ws['B3'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B3'].font = Font(name = 'Arial', size= 8)
        ws['B3'] = 'Fecha Inicio'
        
        ws['C3'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C3'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['C3'].font = Font(name = 'Arial', size= 8)
        ws['C3'].value = fecha_inicio
    
        ws['B4'].alignment = Alignment(horizontal= "left", vertical= "center")
        ws['B4'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B4'].font = Font(name = 'Arial', size= 8)
        ws['B4'] = 'Fecha Fin'
        
        ws['C4'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C4'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['C4'].font = Font(name = 'Arial', size= 8)
        ws['C4'].number_format = "dd-mm-yyyy"
        ws['C4'].value = fecha_fin
        
        # cambia el alto de la columna
        ws.row_dimensions[1].height = 25
        # cambia el ancho de la columna
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 39
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 39
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 39
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 39
        # linea de division
        ws.freeze_panes = 'AL8'

        # crea cabecera
        ws['B6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['B6'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['B6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['B6'] = 'DNI'
        ws.merge_cells('B6:B7')

        ws['C6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['C6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['C6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['C6'] = '1° VISITA'
        ws.merge_cells('C6:D6')
        
        ws['C7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C7'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['C7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['C7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['C7'] = 'FECHA'

        ws['D7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['D7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['D7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['D7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['D7'] = 'ESTABLECIMIENTO'
        ##
        ws['E6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['E6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['E6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['E6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['E6'] = '2° VISITA'
        ws.merge_cells('E6:F6')
        
        ws['E7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['E7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['E7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['E7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['E7'] = 'FECHA'


        ws['F7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['F7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['F7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['F7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['F7'] = 'ESTABLECIMIENTO'

        ws['G6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['G6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['G6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['G6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['G6'] = '3° VISITA'
        ws.merge_cells('G6:H6')
        
        ws['G7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['G7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['G7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['G7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['G7'] = 'FECHA'
        # celda 
        ws['H7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['H7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['H7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['H7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['H7'] = 'ESTABLECIMIENTO'
        # celda 
        ws['I6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['I6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['I6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['I6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['I6'] = '4° VISITA'
        ws.merge_cells('I6:J6')
        
        
        ws['I7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['I7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['I7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['I7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['I7'] = 'FECHA'
        # celda 
        ws['J7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['J7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['J7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['J7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['J7'] = 'ESTABLECIMIENTO'

        # Pintamos los datos del reporte - RED
        cont = 8       
        for q in query:   
            ws.cell(row = cont , column=2).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=2).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=2).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column=2).value = q.Numero_Documento_Paciente

            ws.cell(row = cont , column=3).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=3).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=3).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column=3).value = q.FECHA_VISITA_1
            
            ws.cell(row = cont , column=4).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=4).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=4).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column= 4).value = q.EESS_VISITA_1
            
            ws.cell(row = cont , column=5).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=5).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=5).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column= 5).value = q.FECHA_VISITA_2
            
            ws.cell(row = cont , column=6).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=6).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=6).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column= 6).value = q.EESS_VISITA_2
            
            ws.cell(row = cont , column=7).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=7).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=7).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column= 7).value = q.FECHA_VISITA_3
            
            ws.cell(row = cont , column=8).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=8).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=8).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column= 8).value = q.EESS_VISITA_3
            
            ws.cell(row = cont , column=9).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=9).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=9).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column= 9).value = q.FECHA_VISITA_4
            
            ws.cell(row = cont , column=10).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=10).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=10).font = Font(name = 'Calibri', size= 8)
            ws.cell(row = cont , column= 10).value = q.EESS_VISITA_4
            
            cont+=1

                
        #Respuesta de Django
        #Establecer el nombre del archivo
        nombre_archivo = "rpt_seguimiento_visita_dis.xlsx"
        #Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

#############################################
# ----- RPT DISCAPACIDAD POSTGRES -----------
#############################################
class RptDiscapacidad2(View):
    def get(self, request):
        # Llamada a la función PostgreSQL
        results = self.get_results_from_postgres(2023,1,2,1,6,631)  # Puedes ajustar los valores según tus necesidades
      
        # Renderizar el template y pasar los resultados
        return render(request, 'rpt_discapacidad/rpt_operacional_dis.html', {'results': results})

    def get_results_from_postgres(self, anio, mes_inicio, mes_fin, cod_red, cod_microred, cod_establec):
        # Establecer una conexión a la base de datos    
        with connection.cursor() as cursor:
            # Ejecutar la función PostgreSQL
            cursor.execute(f"SELECT * FROM rpt_discapacidad2({anio}, {mes_inicio}, {mes_fin}, {cod_red}, {cod_microred}, {cod_establec})")         
            # Obtener los resultados
            results = cursor.fetchall()

        return results

################################################
# SITUACION PADRON NOMINAL - VISITA DOMICILARIO
################################################
@login_required
def index(request):
    r_chyo = 'CHANCHAMAYO'
    r_jauja = 'JAUJA'
    r_junin = 'JUNIN'
    r_pki = 'PICHANAKI'
    r_chupaca = 'RED DE SALUD CHUPACA'
    r_pangoa = 'SAN MARTIN DE PANGOA'
    r_satipo = 'SATIPO'
    r_tarma = 'TARMA'
    r_valle = 'VALLE DEL MANTARO'
    
    t_red = RptVisita.objects.values('Red').annotate(
                                                            suma_num=Sum('num'), 
                                                            suma_den=Sum('den')
                                                            ).annotate(
                                                                porcentaje=ExpressionWrapper(
                                                                    (F('suma_num') * 100.0) / F('suma_den'),
                                                                    output_field=FloatField()
                                                                )
                                                            ).order_by('Red')
    
    t_chyo = RptVisita.objects.values('Codigo_Unico','Nombre_Establecimiento').annotate(
                                                            suma_num=Sum('num'), 
                                                            suma_den=Sum('den'),
                                                            suma_v1=Sum('visita1'),
                                                            suma_v2=Sum('visita2'),
                                                            suma_v3=Sum('visita3'),
                                                            suma_v4=Sum('visita4')
                                                            ).annotate(
                                                                porcentaje=ExpressionWrapper(
                                                                    (F('suma_num') * 100.0) / F('suma_den'),
                                                                    output_field=FloatField()
                                                                )
                                                            ).filter(Red=r_chyo).order_by('Nombre_Establecimiento')
                                                            
    t_jauja = RptVisita.objects.values('Codigo_Unico','Nombre_Establecimiento').annotate(
                                                            suma_num=Sum('num'), 
                                                            suma_den=Sum('den'),
                                                            suma_v1=Sum('visita1'),
                                                            suma_v2=Sum('visita2'),
                                                            suma_v3=Sum('visita3'),
                                                            suma_v4=Sum('visita4')
                                                            ).annotate(
                                                                porcentaje=ExpressionWrapper(
                                                                    (F('suma_num') * 100.0) / F('suma_den'),
                                                                    output_field=FloatField()
                                                                )
                                                            ).filter(Red=r_jauja).order_by('Nombre_Establecimiento')

    t_junin = RptVisita.objects.values('Codigo_Unico','Nombre_Establecimiento').annotate(
                                                        suma_num=Sum('num'), 
                                                        suma_den=Sum('den'),
                                                        suma_v1=Sum('visita1'),
                                                        suma_v2=Sum('visita2'),
                                                        suma_v3=Sum('visita3'),
                                                        suma_v4=Sum('visita4')
                                                        ).annotate(
                                                            porcentaje=ExpressionWrapper(
                                                                (F('suma_num') * 100.0) / F('suma_den'),
                                                                output_field=FloatField()
                                                            )
                                                        ).filter(Red=r_junin).order_by('Nombre_Establecimiento')
                                                        
    t_pki = RptVisita.objects.values('Codigo_Unico','Nombre_Establecimiento').annotate(
                                                        suma_num=Sum('num'), 
                                                        suma_den=Sum('den'),
                                                        suma_v1=Sum('visita1'),
                                                        suma_v2=Sum('visita2'),
                                                        suma_v3=Sum('visita3'),
                                                        suma_v4=Sum('visita4')
                                                        ).annotate(
                                                            porcentaje=ExpressionWrapper(
                                                                (F('suma_num') * 100.0) / F('suma_den'),
                                                                output_field=FloatField()
                                                            )
                                                        ).filter(Red=r_pki).order_by('Nombre_Establecimiento')
                                                        
    t_chupaca = RptVisita.objects.values('Codigo_Unico','Nombre_Establecimiento').annotate(
                                                        suma_num=Sum('num'), 
                                                        suma_den=Sum('den'),
                                                        suma_v1=Sum('visita1'),
                                                        suma_v2=Sum('visita2'),
                                                        suma_v3=Sum('visita3'),
                                                        suma_v4=Sum('visita4')
                                                        ).annotate(
                                                            porcentaje=ExpressionWrapper(
                                                                (F('suma_num') * 100.0) / F('suma_den'),
                                                                output_field=FloatField()
                                                            )
                                                        ).filter(Red=r_chupaca).order_by('Nombre_Establecimiento')
                                                        
    t_pangoa = RptVisita.objects.values('Codigo_Unico','Nombre_Establecimiento').annotate(
                                                        suma_num=Sum('num'), 
                                                        suma_den=Sum('den'),
                                                        suma_v1=Sum('visita1'),
                                                        suma_v2=Sum('visita2'),
                                                        suma_v3=Sum('visita3'),
                                                        suma_v4=Sum('visita4')
                                                        ).annotate(
                                                            porcentaje=ExpressionWrapper(
                                                                (F('suma_num') * 100.0) / F('suma_den'),
                                                                output_field=FloatField()
                                                            )
                                                        ).filter(Red=r_pangoa).order_by('Nombre_Establecimiento')
                                                         
    t_satipo = RptVisita.objects.values('Codigo_Unico','Nombre_Establecimiento').annotate(
                                                        suma_num=Sum('num'), 
                                                        suma_den=Sum('den'),
                                                        suma_v1=Sum('visita1'),
                                                        suma_v2=Sum('visita2'),
                                                        suma_v3=Sum('visita3'),
                                                        suma_v4=Sum('visita4')
                                                        ).annotate(
                                                            porcentaje=ExpressionWrapper(
                                                                (F('suma_num') * 100.0) / F('suma_den'),
                                                                output_field=FloatField()
                                                            )
                                                        ).filter(Red=r_satipo).order_by('Nombre_Establecimiento')
                                                        
    t_tarma = RptVisita.objects.values('Codigo_Unico','Nombre_Establecimiento').annotate(
                                                        suma_num=Sum('num'), 
                                                        suma_den=Sum('den'),
                                                        suma_v1=Sum('visita1'),
                                                        suma_v2=Sum('visita2'),
                                                        suma_v3=Sum('visita3'),
                                                        suma_v4=Sum('visita4')
                                                        ).annotate(
                                                            porcentaje=ExpressionWrapper(
                                                                (F('suma_num') * 100.0) / F('suma_den'),
                                                                output_field=FloatField()
                                                            )
                                                        ).filter(Red=r_tarma).order_by('Nombre_Establecimiento')
                                                        
    t_valle = RptVisita.objects.values('Codigo_Unico','Nombre_Establecimiento').annotate(
                                                        suma_num=Sum('num'), 
                                                        suma_den=Sum('den'),
                                                        suma_v1=Sum('visita1'),
                                                        suma_v2=Sum('visita2'),
                                                        suma_v3=Sum('visita3'),
                                                        suma_v4=Sum('visita4')
                                                        ).annotate(
                                                            porcentaje=ExpressionWrapper(
                                                                (F('suma_num') * 100.0) / F('suma_den'),
                                                                output_field=FloatField()
                                                            )
                                                        ).filter(Red=r_valle).order_by('Nombre_Establecimiento')
    

    context = {
                't_red'    : t_red,
                't_chyo'    : t_chyo,
                't_jauja'   : t_jauja,
                't_junin'   : t_junin,
                't_pki'     : t_pki,
                't_chupaca' : t_chupaca,
                't_pangoa'  : t_pangoa,
                't_satipo'  : t_satipo,
                't_tarma'   : t_tarma,
                't_valle'   : t_valle,
              }
    
    return render(request, 'padron/situacion/index.html', context)

################################################
# GRAFICOS
################################################
#---  VISITA GRAFICOS PRINCIPAL ------------------------------------------------
@login_required
def get_chart(_request):
    # Consulta para obtener los datos de ventas
    # Formatear los datos para pasarlos a la plantilla
    colors = ['#5470C6', '#91CC75', '#EE6666'];   
    
    serie=['18060','18230']
    ## counter = 0
    ## while(counter<12):
    ##      serie.append(randrange(100,400))
    ##      counter += 1

    serie2=['1710','832']
    ## counter2 = 0
    ## while(counter2<12):
    ##     serie2.append(randrange(100,400))
    ##     counter2 += 1
        
    serie3=['10.0','4.6']
    ## counter3 = 0
    ## while(counter3<12):
    ##     serie3.append(randrange(0,100))
    ##     counter3 += 1
   
    chart = {
        'tooltip':{
            'show': True,
            'trigger': "axis",
            'triggerOn': "mousemove|click"    
        },
        'legend':{# Nombre para la leyenda
        },       
        'xAxis':[
            {
                'type':"category",
                'data':["ENE","FEB","MAR","ABR","MAY","JUN","JUL","AGO","SET","NOV","DIC"]
            }            
        ],
        'yAxis':[
            {
                'type':"value",
                'axisLine': {
                        'lineStyle': {
                            'color': colors[0]
                        }
                },
            }            
        ],
        'series':[
            {
                'name': 'Meta',
                'data': serie,
                'type': "bar",
            }, 
            {
                'name': 'Avance',
                'data': serie2,
                'type': "bar",
            },
            {
                'name': 'Porcentaje',
                'data': serie3,
                'type': "line",
            },             
        ],
        
        
    }
    
    return JsonResponse(chart)

@login_required
def get_chart_ranking(_request):
    # Formatear los datos para pasarlos a la plantilla
    # Obtener las puntuaciones ordenadas descendientemente
    t_red = RptVisita.objects.values('Red').annotate(
                                                        suma_num=Sum('num'), 
                                                        suma_den=Sum('den')
                                                        ).annotate(
                                                            porcentaje=ExpressionWrapper(
                                                                (F('suma_num') * 100.0) / F('suma_den'),
                                                                output_field=FloatField()
                                                            )
                                                        ).order_by('porcentaje')
                                                        
    # Crear listas separadas para cada variable
    redes = [item['Red'] for item in t_red]
    suma_num = [item['suma_num'] for item in t_red]
    suma_den = [item['suma_den'] for item in t_red]
    porcentajes = [item['porcentaje'] for item in t_red]

    colors = ['#5470C6', '#91CC75', '#EE6666'];   
       
    chart_ranking = {
        'title': {
            'text': 'RANKING POR REDES'
        },
        'tooltip':{
            'show': True,
            'trigger': "axis",
            'triggerOn': "mousemove|click",
            'axisPointer' : {
                'type': 'shadow'
            }
        },
        'legend':{# Nombre para la leyenda
        },
        'grid': {
            'left': '3%',
            'right': '4%',
            'bottom': '3%',
            'containLabel': 'true'
        },       
        'xAxis':[
            {
                'type':"value",
                'boundaryGap':["0","0.01"]
            }            
        ],
        'yAxis':[
            {
                'type':"category",
                'data': redes,
            }            
        ],
        'series':[
            {
                'name': 'Meta',
                'data': suma_den,
                'type': "bar",
            }, 
            {
                'name': 'Avance',
                'data': suma_num,
                'type': "bar",
            },
            {
                'name': 'Porcentaje',
                'data': porcentajes,
                'type': "line",
            },             
        ],
        
        
    }
    return JsonResponse(chart_ranking)
    
################################################
# REPORTE DE SEGUIMIENTO
################################################
#--- PROVINCIAS -------------------------------------------------------------
def get_provincias(request,provincias_id):
    provincias = Provincia.objects.all()
    meses = Item_mes.objects.all()
    context = {
                'provincias': provincias,
                'meses':meses
              }
    return render(request, 'padron/situacion/provincias.html', context)

#--- PROVINCIAS EXCEL -------------------------------------------------------------
class RptProvinciaVistaExcel(TemplateView):
    def get(self,request,*args,**kwargs):
        # variables ingresadas
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin') 
        provincia = request.GET.get('provincia')
        # creacion de la consulta
        query = RptVisita.objects.filter(mes__range=[fecha_inicio, fecha_fin]).filter(ubigeo__startswith=provincia).order_by('Provincia','Distrito','ap_paterno','ap_materno','nom_nino')
        # Convierte mes a numero
        fecha_inicio = int(fecha_inicio)
        fecha_fin = int(fecha_fin)
        
        meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        nombre_mes_inicio = meses[fecha_inicio - 1]
        nombre_mes_fin = meses[fecha_inicio - 1]      

        # print(query)
        # creacion de archivo
        wb = Workbook() #crea libro de trabajo
        ws = wb.active #Primera hoja

        # crea titulo del reporte
        ws['B1'].alignment = Alignment(horizontal= "center", vertical="center")
        ws['B1'].font = Font(name = 'Arial', size= 14, bold = True)
        ws['B1'] = 'REPORTE DE SEGUIMIENTO DE NIÑOS Y NIÑAS MENORES 12 MESES QUE RECIBEN 04 CONSEJERIAS A TRAVES DE VISITA DOMICILIARIA'
        # cambina celdas
        ws.merge_cells('B1:U1')

        ws['B3'].alignment = Alignment(horizontal= "left", vertical= "center")
        ws['B3'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B3'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['B3'] = 'Fecha Inicio'
        
        ws['C3'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C3'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['C3'].font = Font(name = 'Arial', size= 8)
        ws['C3'].value = nombre_mes_inicio
    
        ws['B4'].alignment = Alignment(horizontal= "left", vertical= "center")
        ws['B4'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B4'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['B4'] = 'Fecha Fin'
        
        ws['C4'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C4'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['C4'].font = Font(name = 'Arial', size= 8)
        ws['C4'].number_format = "dd-mm-yyyy"
        ws['C4'].value = nombre_mes_fin
        
        # cambia el alto de la columna
        ws.row_dimensions[1].height = 25
        # cambia el ancho de la columna
        ws.column_dimensions['B'].width = 9
        ws.column_dimensions['C'].width = 13
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 6
        ws.column_dimensions['H'].width = 7
        ws.column_dimensions['I'].width = 25
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 8
        ws.column_dimensions['L'].width = 9
        ws.column_dimensions['M'].width = 9
        ws.column_dimensions['N'].width = 18
        ws.column_dimensions['O'].width = 16
        ws.column_dimensions['P'].width = 16
        ws.column_dimensions['Q'].width = 16
        ws.column_dimensions['R'].width = 20
        ws.column_dimensions['S'].width = 9
        ws.column_dimensions['T'].width = 18
        ws.column_dimensions['U'].width = 10
        ws.column_dimensions['V'].width = 5
        ws.column_dimensions['W'].width = 10
        ws.column_dimensions['X'].width = 5
        ws.column_dimensions['Y'].width = 10
        ws.column_dimensions['Z'].width = 5
        ws.column_dimensions['AA'].width = 9
        ws.column_dimensions['AB'].width = 6
        ws.column_dimensions['AC'].width = 6
        ws.column_dimensions['AD'].width = 6
        # linea de division
        #ws.freeze_panes = 'AL8'

        # crea cabecera
        ws['B6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['B6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['B6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['B6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['B6'] = 'DATOS DEL MENOR'
        ws.merge_cells('B6:N6')
        # crea cabecera 2
        ws['O6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['O6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['O6'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['O6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['O6'] = 'DATOS DE LA CONSEJERIA A TRAVES DE LA VISITA'
        ws.merge_cells('O6:AB6')
        # crea cabecera 2
        ws['AC6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AC6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['AC6'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AC6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AC6'] = 'INDICADOR'
        ws.merge_cells('AC6:AD6')
        
        
        ws['B7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['B7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['B7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['B7'] = 'DNI/CNV'
        
        ws['C7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C7'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['C7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['C7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['C7'] = 'AP PATERNO'

        ws['D7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['D7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['D7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['D7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['D7'] = 'AP MATERNO'
        ##
       
        ws['E7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['E7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['E7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['E7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['E7'] = 'NOMBRES'

        ws['F7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['F7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['F7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['F7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['F7'] = 'FECHA NAC'
        
        ws['G7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['G7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['G7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['G7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['G7'] = 'EDAD'
        # celda 
        ws['H7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['H7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['H7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['H7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['H7'] = 'SEGURO'
        # celda        
        ws['I7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['I7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['I7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['I7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['I7'] = 'DIRECCION'
        # celda 
        ws['J7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['J7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['J7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['J7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['J7'] = 'VISITADO'
        # celda 
        ws['K7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['K7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['K7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['K7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['K7'] = 'ENCONT'
        # celda 
        ws['L7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['L7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['L7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['L7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['L7'] = 'DNI MADRE'
        # celda 
        ws['M7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['M7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['M7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['M7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['M7'] = 'CELULAR'
        # celda 
        ws['N7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['N7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['N7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['N7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['N7'] = 'ENTIDAD PN'
        
        ws['O7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['O7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['O7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['O7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['O7'] = 'PROVINCIA'
        
        ws['P7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['P7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['P7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['P7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['P7'] = 'DISTRITO'
        
        ws['Q7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['Q7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['Q7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['Q7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['Q7'] = 'RED'
        
        ws['R7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['R7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['R7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['R7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['R7'] = 'MICRORED'
        
        ws['S7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['S7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['S7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['S7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['S7'] = 'COD EST'
        
        ws['T7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['T7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['T7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['T7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['T7'] = 'NOMBRE ESTABLEC'
        
        ws['U7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['U7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['U7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['U7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['U7'] = '1RA VISITA'
        
        ws['V7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['V7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['V7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['V7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['V7'] = '1V'
        
        ws['W7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['W7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['W7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['W7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['W7'] = '2DA VISITA'
        
        ws['X7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['X7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['X7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['X7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['X7'] = '2V'
        
        ws['Y7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['Y7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['Y7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['Y7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['Y7'] = '3RA VISITA'
        
        ws['Z7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['Z7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['Z7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['Z7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['Z7'] = '3V'
        
        ws['AA7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AA7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AA7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AA7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AA7'] = '4TA VISITA'
        
        ws['AB7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AB7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AB7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AB7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AB7'] = '4V'
        
        ws['AC7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AC7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AC7'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AC7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AC7'] = 'DEN'
        
        ws['AD7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AD7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AD7'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AD7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AD7'] = 'NUM'
        # Pintamos los datos del reporte - RED
        cont = 8      
        
        for q in query:   
            
            # Iconos
            CHECK_ICON = "✔"
            X_ICON = "✖️"

            # Icono a usar
            icono_visita_1 = CHECK_ICON if q.visita1 == 1 else X_ICON
            icono_visita_2 = CHECK_ICON if q.visita2 == 1 else X_ICON
            icono_visita_3 = CHECK_ICON if q.visita3 == 1 else X_ICON
            icono_visita_4 = CHECK_ICON if q.visita4 == 1 else X_ICON
            
            ws.cell(row = cont , column=2).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=2).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=2).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=2).value = q.num_doc

            ws.cell(row = cont , column=3).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=3).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=3).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=3).value = q.ap_paterno
            
            ws.cell(row = cont , column=4).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=4).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=4).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=4).value = q.ap_materno
            
            ws.cell(row = cont , column=5).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=5).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=5).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=5).value = q.nom_nino
            
            ws.cell(row = cont , column=6).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=6).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=6).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=6).value = q.fecha_nac
            
            ws.cell(row = cont , column=7).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=7).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=7).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=7).value = q.edad_mes
            
            ws.cell(row = cont , column=8).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=8).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=8).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=8).value = q.seguro
            
            ws.cell(row = cont , column=9).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=9).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=9).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=9).value = q.direccion
            
            ws.cell(row = cont , column=10).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=10).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=10).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=10).value = q.visitado
            
            ws.cell(row = cont , column=11).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=11).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=11).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=11).value = q.encontrado
            
            ws.cell(row = cont , column=12).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=12).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=12).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=12).value = q.dni_mama
            
            ws.cell(row = cont , column=13).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=13).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=13).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=13).value = q.num_cel
            
            ws.cell(row = cont , column=14).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=14).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=14).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=14).value = q.pn_reg
            
            ws.cell(row = cont , column=15).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=15).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=15).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=15).value = q.Provincia
            
            ws.cell(row = cont , column=16).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=16).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=16).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=16).value = q.Distrito
            
            ws.cell(row = cont , column=17).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=17).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=17).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=17).value = q.Red
            
            ws.cell(row = cont , column=18).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=18).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=18).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=18).value = q.MicroRed
            
            ws.cell(row = cont , column=19).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=19).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=19).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=19).value = q.Codigo_Unico
            
            ws.cell(row = cont , column=20).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=20).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=20).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=20).value = q.Nombre_Establecimiento
            
            ws.cell(row = cont , column=21).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=21).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=21).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=21).value = q.v_fecha1
            
            ws.cell(row = cont , column=22).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=22).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=22).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=22).value = icono_visita_1
            
            ws.cell(row = cont , column=23).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=23).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=23).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=23).value = q.v_fecha2
            
            ws.cell(row = cont , column=24).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=24).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=24).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=24).value = icono_visita_2
            
            ws.cell(row = cont , column=25).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=25).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=25).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=25).value = q.v_fecha3
            
            ws.cell(row = cont , column=26).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=26).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=26).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=26).value = icono_visita_3
            
            ws.cell(row = cont , column=27).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=27).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=27).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=27).value = q.v_fecha4
            
            ws.cell(row = cont , column=28).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=28).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=28).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=28).value = icono_visita_4
            
            ws.cell(row = cont , column=29).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=29).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=29).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=29).value = q.den
            
            ws.cell(row = cont , column=30).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=30).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=30).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=30).value = q.num
            
            
            cont+=1
               
        #Respuesta de Django
        #Establecer el nombre del archivo
        nombre_archivo = "rpt_seg_visita.xlsx"
        #Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

#--- DISTRITOS -------------------------------------------------------------
def get_distritos(request, distritos_id):
    provincias = Provincia.objects.all()
    meses = Item_mes.objects.all()
    context = {
                'provincias': provincias,
                'meses':meses
              }
    return render(request, 'padron/situacion/distritos.html',context)

def p_distritos(request):
    provincias = request.GET.get('provincia')
    distritos = Distrito.objects.filter(cod_provincia=provincias)
    context= {
            'provincias': provincias,
            'distritos': distritos
             }  
    return render(request, 'padron/situacion/partials/p_distritos.html',context)

#--- DISTRITOS EXCEL -------------------------------------------------------------
class RptDistritoVistaExcel(TemplateView):
    def get(self,request,*args,**kwargs):
        # variables ingresadas
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin') 
        provincia = request.GET.get('provincia')
        distrito = request.GET.get('distritos')
        
        #print(query)
        # creacion de la consulta
        query = RptVisita.objects.filter(mes__range=[fecha_inicio, fecha_fin]).filter(ubigeo=distrito).order_by('ap_paterno','ap_materno','nom_nino','Provincia','Distrito')
        # Convierte mes a numero
        fecha_inicio = int(fecha_inicio)
        fecha_fin = int(fecha_fin)
        
        meses = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SETIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
        nombre_mes_inicio = meses[fecha_inicio - 1]
        nombre_mes_fin = meses[fecha_inicio - 1]      

        # creacion de archivo
        wb = Workbook() #crea libro de trabajo
        ws = wb.active #Primera hoja

        # crea titulo del reporte
        ws['B1'].alignment = Alignment(horizontal= "center", vertical="center")
        ws['B1'].font = Font(name = 'Arial', size= 14, bold = True)
        ws['B1'] = 'REPORTE DE SEGUIMIENTO DE NIÑOS Y NIÑAS MENORES 12 MESES QUE RECIBEN 04 CONSEJERIAS A TRAVES DE VISITA DOMICILIARIA'
        # cambina celdas
        ws.merge_cells('B1:U1')

        ws['B3'].alignment = Alignment(horizontal= "left", vertical= "center")
        ws['B3'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B3'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['B3'] = 'Fecha Inicio'
        
        ws['C3'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C3'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['C3'].font = Font(name = 'Arial', size= 8)
        ws['C3'].value = nombre_mes_inicio
    
        ws['B4'].alignment = Alignment(horizontal= "left", vertical= "center")
        ws['B4'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B4'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['B4'] = 'Fecha Fin'
        
        ws['C4'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C4'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['C4'].font = Font(name = 'Arial', size= 8)
        ws['C4'].number_format = "dd-mm-yyyy"
        ws['C4'].value = nombre_mes_fin
        
        # cambia el alto de la columna
        ws.row_dimensions[1].height = 25
        # cambia el ancho de la columna
        ws.column_dimensions['B'].width = 9
        ws.column_dimensions['C'].width = 13
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 6
        ws.column_dimensions['H'].width = 7
        ws.column_dimensions['I'].width = 25
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 8
        ws.column_dimensions['L'].width = 9
        ws.column_dimensions['M'].width = 9
        ws.column_dimensions['N'].width = 18
        ws.column_dimensions['O'].width = 16
        ws.column_dimensions['P'].width = 16
        ws.column_dimensions['Q'].width = 16
        ws.column_dimensions['R'].width = 20
        ws.column_dimensions['S'].width = 9
        ws.column_dimensions['T'].width = 18
        ws.column_dimensions['U'].width = 10
        ws.column_dimensions['V'].width = 5
        ws.column_dimensions['W'].width = 10
        ws.column_dimensions['X'].width = 5
        ws.column_dimensions['Y'].width = 10
        ws.column_dimensions['Z'].width = 5
        ws.column_dimensions['AA'].width = 9
        ws.column_dimensions['AB'].width = 6
        ws.column_dimensions['AC'].width = 6
        ws.column_dimensions['AD'].width = 6
        # linea de division
        #ws.freeze_panes = 'AL8'

        # crea cabecera
        ws['B6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['B6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['B6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['B6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['B6'] = 'DATOS DEL MENOR'
        ws.merge_cells('B6:N6')
        # crea cabecera 2
        ws['O6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['O6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['O6'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['O6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['O6'] = 'DATOS DE LA CONSEJERIA A TRAVES DE LA VISITA'
        ws.merge_cells('O6:AB6')
        # crea cabecera 3
        ws['AC6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AC6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['AC6'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AC6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AC6'] = 'INDICADOR'
        ws.merge_cells('AC6:AD6')
              
        ws['B7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['B7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['B7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['B7'] = 'DNI/CNV'
        
        ws['C7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C7'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['C7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['C7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['C7'] = 'AP PATERNO'

        ws['D7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['D7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['D7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['D7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['D7'] = 'AP MATERNO'
        ##
       
        ws['E7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['E7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['E7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['E7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['E7'] = 'NOMBRES'

        ws['F7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['F7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['F7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['F7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['F7'] = 'FECHA NAC'
        
        ws['G7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['G7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['G7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['G7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['G7'] = 'EDAD'
        # celda 
        ws['H7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['H7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['H7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['H7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['H7'] = 'SEGURO'
        # celda        
        ws['I7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['I7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['I7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['I7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['I7'] = 'DIRECCION'
        # celda 
        ws['J7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['J7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['J7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['J7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['J7'] = 'VISITADO'
        # celda 
        ws['K7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['K7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['K7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['K7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['K7'] = 'ENCONT'
        # celda 
        ws['L7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['L7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['L7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['L7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['L7'] = 'DNI MADRE'
        # celda 
        ws['M7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['M7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['M7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['M7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['M7'] = 'CELULAR'
        # celda 
        ws['N7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['N7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['N7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['N7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['N7'] = 'ENTIDAD PN'
        
        ws['O7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['O7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['O7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['O7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['O7'] = 'PROVINCIA'
        
        ws['P7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['P7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['P7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['P7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['P7'] = 'DISTRITO'
        
        ws['Q7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['Q7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['Q7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['Q7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['Q7'] = 'RED'
        
        ws['R7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['R7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['R7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['R7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['R7'] = 'MICRORED'
        
        ws['S7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['S7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['S7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['S7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['S7'] = 'COD EST'
        
        ws['T7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['T7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['T7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['T7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['T7'] = 'NOMBRE ESTABLEC'
        
        ws['U7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['U7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['U7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['U7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['U7'] = '1RA VISITA'
        
        ws['V7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['V7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['V7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['V7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['V7'] = '1V'
        
        ws['W7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['W7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['W7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['W7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['W7'] = '2DA VISITA'
        
        ws['X7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['X7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['X7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['X7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['X7'] = '2V'
        
        ws['Y7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['Y7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['Y7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['Y7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['Y7'] = '3RA VISITA'
        
        ws['Z7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['Z7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['Z7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['Z7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['Z7'] = '3V'
        
        ws['AA7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AA7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AA7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AA7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AA7'] = '4TA VISITA'
        
        ws['AB7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AB7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AB7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AB7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AB7'] = '4V'
        
        ws['AC7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AC7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AC7'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AC7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AC7'] = 'DEN'
        
        ws['AD7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AD7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AD7'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AD7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AD7'] = 'NUM'
        # Pintamos los datos del reporte - RED
        cont = 8      
        
        for q in query:   
            
            # Iconos
            CHECK_ICON = "✔"
            X_ICON = "✖️"

            # Icono a usar
            icono_visita_1 = CHECK_ICON if q.visita1 == 1 else X_ICON
            icono_visita_2 = CHECK_ICON if q.visita2 == 1 else X_ICON
            icono_visita_3 = CHECK_ICON if q.visita3 == 1 else X_ICON
            icono_visita_4 = CHECK_ICON if q.visita4 == 1 else X_ICON
            
            ws.cell(row = cont , column=2).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=2).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=2).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=2).value = q.num_doc

            ws.cell(row = cont , column=3).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=3).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=3).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=3).value = q.ap_paterno
            
            ws.cell(row = cont , column=4).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=4).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=4).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=4).value = q.ap_materno
            
            ws.cell(row = cont , column=5).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=5).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=5).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=5).value = q.nom_nino
            
            ws.cell(row = cont , column=6).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=6).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=6).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=6).value = q.fecha_nac
            
            ws.cell(row = cont , column=7).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=7).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=7).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=7).value = q.edad_mes
            
            ws.cell(row = cont , column=8).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=8).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=8).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=8).value = q.seguro
            
            ws.cell(row = cont , column=9).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=9).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=9).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=9).value = q.direccion
            
            ws.cell(row = cont , column=10).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=10).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=10).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=10).value = q.visitado
            
            ws.cell(row = cont , column=11).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=11).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=11).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=11).value = q.encontrado
            
            ws.cell(row = cont , column=12).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=12).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=12).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=12).value = q.dni_mama
            
            ws.cell(row = cont , column=13).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=13).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=13).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=13).value = q.num_cel
            
            ws.cell(row = cont , column=14).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=14).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=14).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=14).value = q.pn_reg
            
            ws.cell(row = cont , column=15).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=15).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=15).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=15).value = q.Provincia
            
            ws.cell(row = cont , column=16).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=16).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=16).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=16).value = q.Distrito
            
            ws.cell(row = cont , column=17).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=17).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=17).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=17).value = q.Red
            
            ws.cell(row = cont , column=18).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=18).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=18).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=18).value = q.MicroRed
            
            ws.cell(row = cont , column=19).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=19).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=19).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=19).value = q.Codigo_Unico
            
            ws.cell(row = cont , column=20).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=20).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=20).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=20).value = q.Nombre_Establecimiento
            
            ws.cell(row = cont , column=21).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=21).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=21).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=21).value = q.v_fecha1
            
            ws.cell(row = cont , column=22).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=22).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=22).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=22).value = icono_visita_1
            
            ws.cell(row = cont , column=23).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=23).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=23).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=23).value = q.v_fecha2
            
            ws.cell(row = cont , column=24).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=24).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=24).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=24).value = icono_visita_2
            
            ws.cell(row = cont , column=25).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=25).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=25).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=25).value = q.v_fecha3
            
            ws.cell(row = cont , column=26).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=26).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=26).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=26).value = icono_visita_3
            
            ws.cell(row = cont , column=27).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=27).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=27).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=27).value = q.v_fecha4
            
            ws.cell(row = cont , column=28).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=28).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=28).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=28).value = icono_visita_4
            
            ws.cell(row = cont , column=29).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=29).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=29).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=29).value = q.den
            
            ws.cell(row = cont , column=30).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=30).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=30).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=30).value = q.num
            
            
            cont+=1
               
        #Respuesta de Django
        #Establecer el nombre del archivo
        nombre_archivo = "rpt_seg_visita.xlsx"
        #Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

#--- REDES ------------------------------------------------------------------
def get_redes(request,redes_id):
    redes = Red.objects.all()
    meses = Item_mes.objects.all()
    context = {
                'redes': redes,
                'meses':meses
              }
    return render(request, 'padron/situacion/redes.html',context)

#--- REDES EXCEL -------------------------------------------------------------
class RptRedVistaExcel(TemplateView):
    def get(self,request,*args,**kwargs):
        # variables ingresadas
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin') 
        red = request.GET.get('red')
        # creacion de la consulta
        query = RptVisita.objects.filter(mes__range=[fecha_inicio, fecha_fin]).filter(Codigo_Red=red).order_by('Red','MicroRed','Nombre_Establecimiento','ap_paterno','ap_materno','nom_nino')
        # Convierte mes a numero
        fecha_inicio = int(fecha_inicio)
        fecha_fin = int(fecha_fin)
        
        meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        nombre_mes_inicio = meses[fecha_inicio - 1]
        nombre_mes_fin = meses[fecha_inicio - 1]      

        # print(query)
        # creacion de archivo
        wb = Workbook() #crea libro de trabajo
        ws = wb.active #Primera hoja

        # crea titulo del reporte
        ws['B1'].alignment = Alignment(horizontal= "center", vertical="center")
        ws['B1'].font = Font(name = 'Arial', size= 14, bold = True)
        ws['B1'] = 'REPORTE DE SEGUIMIENTO DE NIÑOS Y NIÑAS MENORES 12 MESES QUE RECIBEN 04 CONSEJERIAS A TRAVES DE VISITA DOMICILIARIA'
        # cambina celdas
        ws.merge_cells('B1:U1')

        ws['B3'].alignment = Alignment(horizontal= "left", vertical= "center")
        ws['B3'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B3'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['B3'] = 'Fecha Inicio'
        
        ws['C3'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C3'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['C3'].font = Font(name = 'Arial', size= 8)
        ws['C3'].value = nombre_mes_inicio
    
        ws['B4'].alignment = Alignment(horizontal= "left", vertical= "center")
        ws['B4'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B4'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['B4'] = 'Fecha Fin'
        
        ws['C4'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C4'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['C4'].font = Font(name = 'Arial', size= 8)
        ws['C4'].number_format = "dd-mm-yyyy"
        ws['C4'].value = nombre_mes_fin
        
        # cambia el alto de la columna
        ws.row_dimensions[1].height = 25
        # cambia el ancho de la columna
        ws.column_dimensions['B'].width = 9
        ws.column_dimensions['C'].width = 13
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 6
        ws.column_dimensions['H'].width = 7
        ws.column_dimensions['I'].width = 25
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 8
        ws.column_dimensions['L'].width = 9
        ws.column_dimensions['M'].width = 9
        ws.column_dimensions['N'].width = 18
        ws.column_dimensions['O'].width = 16
        ws.column_dimensions['P'].width = 16
        ws.column_dimensions['Q'].width = 16
        ws.column_dimensions['R'].width = 20
        ws.column_dimensions['S'].width = 9
        ws.column_dimensions['T'].width = 18
        ws.column_dimensions['U'].width = 10
        ws.column_dimensions['V'].width = 5
        ws.column_dimensions['W'].width = 10
        ws.column_dimensions['X'].width = 5
        ws.column_dimensions['Y'].width = 10
        ws.column_dimensions['Z'].width = 5
        ws.column_dimensions['AA'].width = 9
        ws.column_dimensions['AB'].width = 6
        ws.column_dimensions['AC'].width = 6
        ws.column_dimensions['AD'].width = 6
        # linea de division
        #ws.freeze_panes = 'AL8'

        # crea cabecera
        ws['B6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['B6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['B6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['B6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['B6'] = 'DATOS DEL MENOR'
        ws.merge_cells('B6:N6')
        # crea cabecera 2
        ws['O6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['O6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['O6'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['O6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['O6'] = 'DATOS DE LA CONSEJERIA A TRAVES DE LA VISITA'
        ws.merge_cells('O6:AB6')
        # crea cabecera 2
        ws['AC6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AC6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['AC6'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AC6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AC6'] = 'INDICADOR'
        ws.merge_cells('AC6:AD6')
        
        
        ws['B7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['B7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['B7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['B7'] = 'DNI/CNV'
        
        ws['C7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C7'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['C7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['C7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['C7'] = 'AP PATERNO'

        ws['D7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['D7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['D7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['D7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['D7'] = 'AP MATERNO'
        ##
       
        ws['E7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['E7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['E7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['E7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['E7'] = 'NOMBRES'

        ws['F7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['F7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['F7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['F7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['F7'] = 'FECHA NAC'
        
        ws['G7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['G7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['G7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['G7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['G7'] = 'EDAD'
        # celda 
        ws['H7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['H7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['H7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['H7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['H7'] = 'SEGURO'
        # celda        
        ws['I7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['I7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['I7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['I7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['I7'] = 'DIRECCION'
        # celda 
        ws['J7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['J7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['J7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['J7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['J7'] = 'VISITADO'
        # celda 
        ws['K7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['K7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['K7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['K7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['K7'] = 'ENCONT'
        # celda 
        ws['L7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['L7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['L7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['L7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['L7'] = 'DNI MADRE'
        # celda 
        ws['M7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['M7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['M7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['M7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['M7'] = 'CELULAR'
        # celda 
        ws['N7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['N7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['N7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['N7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['N7'] = 'ENTIDAD PN'
        
        ws['O7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['O7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['O7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['O7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['O7'] = 'PROVINCIA'
        
        ws['P7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['P7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['P7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['P7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['P7'] = 'DISTRITO'
        
        ws['Q7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['Q7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['Q7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['Q7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['Q7'] = 'RED'
        
        ws['R7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['R7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['R7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['R7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['R7'] = 'MICRORED'
        
        ws['S7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['S7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['S7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['S7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['S7'] = 'COD EST'
        
        ws['T7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['T7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['T7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['T7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['T7'] = 'NOMBRE ESTABLEC'
        
        ws['U7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['U7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['U7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['U7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['U7'] = '1RA VISITA'
        
        ws['V7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['V7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['V7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['V7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['V7'] = '1V'
        
        ws['W7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['W7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['W7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['W7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['W7'] = '2DA VISITA'
        
        ws['X7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['X7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['X7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['X7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['X7'] = '2V'
        
        ws['Y7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['Y7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['Y7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['Y7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['Y7'] = '3RA VISITA'
        
        ws['Z7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['Z7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['Z7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['Z7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['Z7'] = '3V'
        
        ws['AA7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AA7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AA7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AA7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AA7'] = '4TA VISITA'
        
        ws['AB7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AB7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AB7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AB7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AB7'] = '4V'
        
        ws['AC7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AC7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AC7'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AC7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AC7'] = 'DEN'
        
        ws['AD7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AD7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AD7'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AD7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AD7'] = 'NUM'
        # Pintamos los datos del reporte - RED
        cont = 8      
        
        for q in query:   
            
            # Iconos
            CHECK_ICON = "✔"
            X_ICON = "✖️"

            # Icono a usar
            icono_visita_1 = CHECK_ICON if q.visita1 == 1 else X_ICON
            icono_visita_2 = CHECK_ICON if q.visita2 == 1 else X_ICON
            icono_visita_3 = CHECK_ICON if q.visita3 == 1 else X_ICON
            icono_visita_4 = CHECK_ICON if q.visita4 == 1 else X_ICON
            
            ws.cell(row = cont , column=2).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=2).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=2).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=2).value = q.num_doc

            ws.cell(row = cont , column=3).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=3).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=3).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=3).value = q.ap_paterno
            
            ws.cell(row = cont , column=4).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=4).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=4).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=4).value = q.ap_materno
            
            ws.cell(row = cont , column=5).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=5).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=5).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=5).value = q.nom_nino
            
            ws.cell(row = cont , column=6).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=6).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=6).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=6).value = q.fecha_nac
            
            ws.cell(row = cont , column=7).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=7).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=7).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=7).value = q.edad_mes
            
            ws.cell(row = cont , column=8).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=8).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=8).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=8).value = q.seguro
            
            ws.cell(row = cont , column=9).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=9).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=9).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=9).value = q.direccion
            
            ws.cell(row = cont , column=10).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=10).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=10).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=10).value = q.visitado
            
            ws.cell(row = cont , column=11).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=11).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=11).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=11).value = q.encontrado
            
            ws.cell(row = cont , column=12).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=12).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=12).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=12).value = q.dni_mama
            
            ws.cell(row = cont , column=13).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=13).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=13).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=13).value = q.num_cel
            
            ws.cell(row = cont , column=14).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=14).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=14).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=14).value = q.pn_reg
            
            ws.cell(row = cont , column=15).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=15).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=15).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=15).value = q.Provincia
            
            ws.cell(row = cont , column=16).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=16).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=16).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=16).value = q.Distrito
            
            ws.cell(row = cont , column=17).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=17).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=17).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=17).value = q.Red
            
            ws.cell(row = cont , column=18).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=18).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=18).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=18).value = q.MicroRed
            
            ws.cell(row = cont , column=19).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=19).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=19).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=19).value = q.Codigo_Unico
            
            ws.cell(row = cont , column=20).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=20).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=20).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=20).value = q.Nombre_Establecimiento
            
            ws.cell(row = cont , column=21).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=21).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=21).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=21).value = q.v_fecha1
            
            ws.cell(row = cont , column=22).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=22).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=22).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=22).value = icono_visita_1
            
            ws.cell(row = cont , column=23).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=23).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=23).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=23).value = q.v_fecha2
            
            ws.cell(row = cont , column=24).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=24).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=24).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=24).value = icono_visita_2
            
            ws.cell(row = cont , column=25).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=25).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=25).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=25).value = q.v_fecha3
            
            ws.cell(row = cont , column=26).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=26).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=26).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=26).value = icono_visita_3
            
            ws.cell(row = cont , column=27).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=27).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=27).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=27).value = q.v_fecha4
            
            ws.cell(row = cont , column=28).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=28).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=28).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=28).value = icono_visita_4
            
            ws.cell(row = cont , column=29).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=29).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=29).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=29).value = q.den
            
            ws.cell(row = cont , column=30).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=30).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=30).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=30).value = q.num
            
            
            cont+=1
               
        #Respuesta de Django
        #Establecer el nombre del archivo
        nombre_archivo = "rpt_seg_visita.xlsx"
        #Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


#--- MICROREDES -------------------------------------------------------------
def get_microredes(request, microredes_id):
    redes = Red.objects.all()
    meses = Item_mes.objects.all()
    context= {
                'redes': redes,
                'meses': meses,
             }
    return render(request, 'padron/situacion/microredes.html', context)

def p_microredes_principal(request):
    redes = request.GET.get('redes')
    microredes = Microred.objects.filter(cod_red=redes)
    context= {
            'microredes': microredes,
             }
    return render(request, 'padron/situacion/partials/p_microredes_principal.html', context)

def p_microredes(request):
    redes = request.GET.get('redes')
    microredes = Microred.objects.filter(cod_red=redes)
    context= {
            'microredes': microredes,
            'is_htmx': True
             }
    return render(request, 'padron/situacion/partials/p_microredes.html', context)

#--- MICROREDES EXCEL -------------------------------------------------------------
class RptMicroredVistaExcel(TemplateView):
    def get(self,request,*args,**kwargs):
        # variables ingresadas
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin') 
        red = request.GET.get('redes')
        microred = request.GET.get('microredes')
        # creacion de la consulta
        query = RptVisita.objects.filter(mes__range=[fecha_inicio, fecha_fin]).filter(Codigo_Red=red).filter(Codigo_MicroRed=microred).order_by('Red','MicroRed','Nombre_Establecimiento','ap_paterno','ap_materno','nom_nino')
        # Convierte mes a numero
        fecha_inicio = int(fecha_inicio)
        fecha_fin = int(fecha_fin)
        
        meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        nombre_mes_inicio = meses[fecha_inicio - 1]
        nombre_mes_fin = meses[fecha_inicio - 1]      

        print(query)
        # creacion de archivo
        wb = Workbook() #crea libro de trabajo
        ws = wb.active #Primera hoja

        # crea titulo del reporte
        ws['B1'].alignment = Alignment(horizontal= "center", vertical="center")
        ws['B1'].font = Font(name = 'Arial', size= 14, bold = True)
        ws['B1'] = 'REPORTE DE SEGUIMIENTO DE NIÑOS Y NIÑAS MENORES 12 MESES QUE RECIBEN 04 CONSEJERIAS A TRAVES DE VISITA DOMICILIARIA'
        # cambina celdas
        ws.merge_cells('B1:U1')

        ws['B3'].alignment = Alignment(horizontal= "left", vertical= "center")
        ws['B3'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B3'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['B3'] = 'Fecha Inicio'
        
        ws['C3'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C3'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['C3'].font = Font(name = 'Arial', size= 8)
        ws['C3'].value = nombre_mes_inicio
    
        ws['B4'].alignment = Alignment(horizontal= "left", vertical= "center")
        ws['B4'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B4'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['B4'] = 'Fecha Fin'
        
        ws['C4'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C4'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['C4'].font = Font(name = 'Arial', size= 8)
        ws['C4'].number_format = "dd-mm-yyyy"
        ws['C4'].value = nombre_mes_fin
        
        # cambia el alto de la columna
        ws.row_dimensions[1].height = 25
        # cambia el ancho de la columna
        ws.column_dimensions['B'].width = 9
        ws.column_dimensions['C'].width = 13
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 6
        ws.column_dimensions['H'].width = 7
        ws.column_dimensions['I'].width = 25
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 8
        ws.column_dimensions['L'].width = 9
        ws.column_dimensions['M'].width = 9
        ws.column_dimensions['N'].width = 18
        ws.column_dimensions['O'].width = 16
        ws.column_dimensions['P'].width = 16
        ws.column_dimensions['Q'].width = 16
        ws.column_dimensions['R'].width = 20
        ws.column_dimensions['S'].width = 9
        ws.column_dimensions['T'].width = 18
        ws.column_dimensions['U'].width = 10
        ws.column_dimensions['V'].width = 5
        ws.column_dimensions['W'].width = 10
        ws.column_dimensions['X'].width = 5
        ws.column_dimensions['Y'].width = 10
        ws.column_dimensions['Z'].width = 5
        ws.column_dimensions['AA'].width = 9
        ws.column_dimensions['AB'].width = 6
        ws.column_dimensions['AC'].width = 6
        ws.column_dimensions['AD'].width = 6
        # linea de division
        #ws.freeze_panes = 'AL8'

        # crea cabecera
        ws['B6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['B6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['B6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['B6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['B6'] = 'DATOS DEL MENOR'
        ws.merge_cells('B6:N6')
        # crea cabecera 2
        ws['O6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['O6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['O6'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['O6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['O6'] = 'DATOS DE LA CONSEJERIA A TRAVES DE LA VISITA'
        ws.merge_cells('O6:AB6')
        # crea cabecera 2
        ws['AC6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AC6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['AC6'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AC6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AC6'] = 'INDICADOR'
        ws.merge_cells('AC6:AD6')
        
        
        ws['B7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['B7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['B7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['B7'] = 'DNI/CNV'
        
        ws['C7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C7'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['C7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['C7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['C7'] = 'AP PATERNO'

        ws['D7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['D7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['D7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['D7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['D7'] = 'AP MATERNO'
        ##
       
        ws['E7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['E7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['E7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['E7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['E7'] = 'NOMBRES'

        ws['F7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['F7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['F7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['F7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['F7'] = 'FECHA NAC'
        
        ws['G7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['G7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['G7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['G7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['G7'] = 'EDAD'
        # celda 
        ws['H7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['H7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['H7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['H7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['H7'] = 'SEGURO'
        # celda        
        ws['I7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['I7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['I7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['I7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['I7'] = 'DIRECCION'
        # celda 
        ws['J7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['J7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['J7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['J7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['J7'] = 'VISITADO'
        # celda 
        ws['K7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['K7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['K7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['K7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['K7'] = 'ENCONT'
        # celda 
        ws['L7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['L7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['L7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['L7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['L7'] = 'DNI MADRE'
        # celda 
        ws['M7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['M7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['M7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['M7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['M7'] = 'CELULAR'
        # celda 
        ws['N7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['N7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['N7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['N7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['N7'] = 'ENTIDAD PN'
        
        ws['O7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['O7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['O7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['O7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['O7'] = 'PROVINCIA'
        
        ws['P7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['P7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['P7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['P7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['P7'] = 'DISTRITO'
        
        ws['Q7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['Q7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['Q7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['Q7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['Q7'] = 'RED'
        
        ws['R7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['R7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['R7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['R7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['R7'] = 'MICRORED'
        
        ws['S7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['S7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['S7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['S7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['S7'] = 'COD EST'
        
        ws['T7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['T7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['T7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['T7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['T7'] = 'NOMBRE ESTABLEC'
        
        ws['U7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['U7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['U7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['U7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['U7'] = '1RA VISITA'
        
        ws['V7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['V7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['V7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['V7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['V7'] = '1V'
        
        ws['W7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['W7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['W7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['W7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['W7'] = '2DA VISITA'
        
        ws['X7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['X7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['X7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['X7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['X7'] = '2V'
        
        ws['Y7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['Y7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['Y7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['Y7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['Y7'] = '3RA VISITA'
        
        ws['Z7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['Z7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['Z7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['Z7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['Z7'] = '3V'
        
        ws['AA7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AA7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AA7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AA7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AA7'] = '4TA VISITA'
        
        ws['AB7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AB7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AB7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AB7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AB7'] = '4V'
        
        ws['AC7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AC7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AC7'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AC7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AC7'] = 'DEN'
        
        ws['AD7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AD7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AD7'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AD7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AD7'] = 'NUM'
        # Pintamos los datos del reporte - RED
        cont = 8      
        
        for q in query:   
            
            # Iconos
            CHECK_ICON = "✔"
            X_ICON = "✖️"

            # Icono a usar
            icono_visita_1 = CHECK_ICON if q.visita1 == 1 else X_ICON
            icono_visita_2 = CHECK_ICON if q.visita2 == 1 else X_ICON
            icono_visita_3 = CHECK_ICON if q.visita3 == 1 else X_ICON
            icono_visita_4 = CHECK_ICON if q.visita4 == 1 else X_ICON
            
            ws.cell(row = cont , column=2).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=2).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=2).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=2).value = q.num_doc

            ws.cell(row = cont , column=3).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=3).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=3).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=3).value = q.ap_paterno
            
            ws.cell(row = cont , column=4).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=4).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=4).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=4).value = q.ap_materno
            
            ws.cell(row = cont , column=5).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=5).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=5).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=5).value = q.nom_nino
            
            ws.cell(row = cont , column=6).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=6).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=6).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=6).value = q.fecha_nac
            
            ws.cell(row = cont , column=7).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=7).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=7).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=7).value = q.edad_mes
            
            ws.cell(row = cont , column=8).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=8).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=8).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=8).value = q.seguro
            
            ws.cell(row = cont , column=9).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=9).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=9).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=9).value = q.direccion
            
            ws.cell(row = cont , column=10).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=10).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=10).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=10).value = q.visitado
            
            ws.cell(row = cont , column=11).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=11).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=11).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=11).value = q.encontrado
            
            ws.cell(row = cont , column=12).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=12).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=12).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=12).value = q.dni_mama
            
            ws.cell(row = cont , column=13).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=13).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=13).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=13).value = q.num_cel
            
            ws.cell(row = cont , column=14).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=14).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=14).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=14).value = q.pn_reg
            
            ws.cell(row = cont , column=15).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=15).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=15).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=15).value = q.Provincia
            
            ws.cell(row = cont , column=16).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=16).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=16).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=16).value = q.Distrito
            
            ws.cell(row = cont , column=17).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=17).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=17).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=17).value = q.Red
            
            ws.cell(row = cont , column=18).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=18).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=18).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=18).value = q.MicroRed
            
            ws.cell(row = cont , column=19).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=19).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=19).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=19).value = q.Codigo_Unico
            
            ws.cell(row = cont , column=20).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=20).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=20).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=20).value = q.Nombre_Establecimiento
            
            ws.cell(row = cont , column=21).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=21).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=21).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=21).value = q.v_fecha1
            
            ws.cell(row = cont , column=22).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=22).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=22).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=22).value = icono_visita_1
            
            ws.cell(row = cont , column=23).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=23).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=23).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=23).value = q.v_fecha2
            
            ws.cell(row = cont , column=24).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=24).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=24).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=24).value = icono_visita_2
            
            ws.cell(row = cont , column=25).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=25).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=25).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=25).value = q.v_fecha3
            
            ws.cell(row = cont , column=26).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=26).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=26).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=26).value = icono_visita_3
            
            ws.cell(row = cont , column=27).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=27).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=27).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=27).value = q.v_fecha4
            
            ws.cell(row = cont , column=28).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=28).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=28).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=28).value = icono_visita_4
            
            ws.cell(row = cont , column=29).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=29).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=29).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=29).value = q.den
            
            ws.cell(row = cont , column=30).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=30).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=30).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=30).value = q.num
            
            
            cont+=1
               
        #Respuesta de Django
        #Establecer el nombre del archivo
        nombre_archivo = "rpt_seg_visita.xlsx"
        #Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


#--- ESTABLECIMIENTOS -------------------------------------------------------
def get_establecimientos(request,establecimiento_id):
    redes = Red.objects.all()
    meses = Item_mes.objects.all()
    context = {
                'redes': redes,
                'meses': meses,
              }
    return render(request, 'padron/situacion/establecimientos.html',context)

def p_establecimientos(request):
    microredes = request.GET.get('microredes')
    establecimientos = Establecimiento.objects.filter(red_microred=microredes)
    context= {
            'establecimientos': establecimientos
             }
    return render(request, 'padron/situacion/partials/p_establecimientos.html', context)

#--- MICROREDES EXCEL -------------------------------------------------------------
class RptEstablecimientoVistaExcel(TemplateView):
    def get(self,request,*args,**kwargs):
        # variables ingresadas
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin') 
        red = request.GET.get('redes')
        microred = request.GET.get('microredes')
        establecimiento = request.GET.get('p_establecimiento')
        
        print(establecimiento)
        # creacion de la consulta
        query = RptVisita.objects.filter(mes__range=[fecha_inicio, fecha_fin]).filter(Id_Establecimiento=establecimiento).order_by('Red','MicroRed','Nombre_Establecimiento','ap_paterno','ap_materno','nom_nino')
        # Convierte mes a numero
        fecha_inicio = int(fecha_inicio)
        fecha_fin = int(fecha_fin)
        
        meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        nombre_mes_inicio = meses[fecha_inicio - 1]
        nombre_mes_fin = meses[fecha_inicio - 1]      

        print(query)
        # creacion de archivo
        wb = Workbook() #crea libro de trabajo
        ws = wb.active #Primera hoja

        # crea titulo del reporte
        ws['B1'].alignment = Alignment(horizontal= "center", vertical="center")
        ws['B1'].font = Font(name = 'Arial', size= 14, bold = True)
        ws['B1'] = 'REPORTE DE SEGUIMIENTO DE NIÑOS Y NIÑAS MENORES 12 MESES QUE RECIBEN 04 CONSEJERIAS A TRAVES DE VISITA DOMICILIARIA'
        # cambina celdas
        ws.merge_cells('B1:U1')

        ws['B3'].alignment = Alignment(horizontal= "left", vertical= "center")
        ws['B3'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B3'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['B3'] = 'Fecha Inicio'
        
        ws['C3'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C3'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['C3'].font = Font(name = 'Arial', size= 8)
        ws['C3'].value = nombre_mes_inicio
    
        ws['B4'].alignment = Alignment(horizontal= "left", vertical= "center")
        ws['B4'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B4'].font = Font(name = 'Arial', size= 8, bold = True)
        ws['B4'] = 'Fecha Fin'
        
        ws['C4'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C4'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['C4'].font = Font(name = 'Arial', size= 8)
        ws['C4'].number_format = "dd-mm-yyyy"
        ws['C4'].value = nombre_mes_fin
        
        # cambia el alto de la columna
        ws.row_dimensions[1].height = 25
        # cambia el ancho de la columna
        ws.column_dimensions['B'].width = 9
        ws.column_dimensions['C'].width = 13
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 6
        ws.column_dimensions['H'].width = 7
        ws.column_dimensions['I'].width = 25
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 8
        ws.column_dimensions['L'].width = 9
        ws.column_dimensions['M'].width = 9
        ws.column_dimensions['N'].width = 18
        ws.column_dimensions['O'].width = 16
        ws.column_dimensions['P'].width = 16
        ws.column_dimensions['Q'].width = 16
        ws.column_dimensions['R'].width = 20
        ws.column_dimensions['S'].width = 9
        ws.column_dimensions['T'].width = 18
        ws.column_dimensions['U'].width = 10
        ws.column_dimensions['V'].width = 5
        ws.column_dimensions['W'].width = 10
        ws.column_dimensions['X'].width = 5
        ws.column_dimensions['Y'].width = 10
        ws.column_dimensions['Z'].width = 5
        ws.column_dimensions['AA'].width = 9
        ws.column_dimensions['AB'].width = 6
        ws.column_dimensions['AC'].width = 6
        ws.column_dimensions['AD'].width = 6
        # linea de division
        #ws.freeze_panes = 'AL8'

        # crea cabecera
        ws['B6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['B6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['B6'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['B6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['B6'] = 'DATOS DEL MENOR'
        ws.merge_cells('B6:N6')
        # crea cabecera 2
        ws['O6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['O6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['O6'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['O6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['O6'] = 'DATOS DE LA CONSEJERIA A TRAVES DE LA VISITA'
        ws.merge_cells('O6:AB6')
        # crea cabecera 2
        ws['AC6'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AC6'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['AC6'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AC6'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AC6'] = 'INDICADOR'
        ws.merge_cells('AC6:AD6')
        
        
        ws['B7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['B7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['B7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['B7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['B7'] = 'DNI/CNV'
        
        ws['C7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['C7'].border = Border(left = Side(border_style = "thin"), 
                                     right = Side(border_style = "thin"), 
                                     top = Side(border_style = "thin"), 
                                     bottom = Side(border_style = "thin"))
        ws['C7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['C7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['C7'] = 'AP PATERNO'

        ws['D7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['D7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['D7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['D7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['D7'] = 'AP MATERNO'
        ##
       
        ws['E7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['E7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['E7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['E7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['E7'] = 'NOMBRES'

        ws['F7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['F7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['F7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['F7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['F7'] = 'FECHA NAC'
        
        ws['G7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['G7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['G7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['G7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['G7'] = 'EDAD'
        # celda 
        ws['H7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['H7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['H7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['H7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['H7'] = 'SEGURO'
        # celda        
        ws['I7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['I7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['I7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['I7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['I7'] = 'DIRECCION'
        # celda 
        ws['J7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['J7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['J7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['J7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['J7'] = 'VISITADO'
        # celda 
        ws['K7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['K7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['K7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['K7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['K7'] = 'ENCONT'
        # celda 
        ws['L7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['L7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['L7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['L7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['L7'] = 'DNI MADRE'
        # celda 
        ws['M7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['M7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['M7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['M7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['M7'] = 'CELULAR'
        # celda 
        ws['N7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['N7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['N7'].fill = PatternFill(start_color = 'DDF2FD', end_color='DDF2FD', fill_type="solid")
        ws['N7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['N7'] = 'ENTIDAD PN'
        
        ws['O7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['O7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['O7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['O7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['O7'] = 'PROVINCIA'
        
        ws['P7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['P7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['P7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['P7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['P7'] = 'DISTRITO'
        
        ws['Q7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['Q7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['Q7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['Q7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['Q7'] = 'RED'
        
        ws['R7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['R7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['R7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['R7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['R7'] = 'MICRORED'
        
        ws['S7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['S7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['S7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['S7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['S7'] = 'COD EST'
        
        ws['T7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['T7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['T7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['T7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['T7'] = 'NOMBRE ESTABLEC'
        
        ws['U7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['U7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['U7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['U7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['U7'] = '1RA VISITA'
        
        ws['V7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['V7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['V7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['V7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['V7'] = '1V'
        
        ws['W7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['W7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['W7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['W7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['W7'] = '2DA VISITA'
        
        ws['X7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['X7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['X7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['X7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['X7'] = '2V'
        
        ws['Y7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['Y7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['Y7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['Y7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['Y7'] = '3RA VISITA'
        
        ws['Z7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['Z7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['Z7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['Z7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['Z7'] = '3V'
        
        ws['AA7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AA7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AA7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AA7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AA7'] = '4TA VISITA'
        
        ws['AB7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AB7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AB7'].fill = PatternFill(start_color = 'FDEDDD', end_color='FDEDDD', fill_type="solid")
        ws['AB7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AB7'] = '4V'
        
        ws['AC7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AC7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AC7'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AC7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AC7'] = 'DEN'
        
        ws['AD7'].alignment = Alignment(horizontal= "center", vertical= "center")
        ws['AD7'].border = Border(left = Side(border_style = "thin"), 
                                 right = Side(border_style = "thin"), 
                                 top = Side(border_style = "thin"), 
                                 bottom = Side(border_style = "thin"))
        ws['AD7'].fill = PatternFill(start_color = 'DDFDED', end_color='DDFDED', fill_type="solid")
        ws['AD7'].font = Font(name = 'Arial', size= 9, bold = True)
        ws['AD7'] = 'NUM'
        # Pintamos los datos del reporte - RED
        cont = 8      
        
        for q in query:   
            
            # Iconos
            CHECK_ICON = "✔"
            X_ICON = "✖️"

            # Icono a usar
            icono_visita_1 = CHECK_ICON if q.visita1 == 1 else X_ICON
            icono_visita_2 = CHECK_ICON if q.visita2 == 1 else X_ICON
            icono_visita_3 = CHECK_ICON if q.visita3 == 1 else X_ICON
            icono_visita_4 = CHECK_ICON if q.visita4 == 1 else X_ICON
            
            ws.cell(row = cont , column=2).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=2).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=2).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=2).value = q.num_doc

            ws.cell(row = cont , column=3).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=3).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=3).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=3).value = q.ap_paterno
            
            ws.cell(row = cont , column=4).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=4).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=4).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=4).value = q.ap_materno
            
            ws.cell(row = cont , column=5).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=5).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=5).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=5).value = q.nom_nino
            
            ws.cell(row = cont , column=6).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=6).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=6).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=6).value = q.fecha_nac
            
            ws.cell(row = cont , column=7).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=7).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=7).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=7).value = q.edad_mes
            
            ws.cell(row = cont , column=8).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=8).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=8).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=8).value = q.seguro
            
            ws.cell(row = cont , column=9).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=9).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=9).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=9).value = q.direccion
            
            ws.cell(row = cont , column=10).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=10).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=10).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=10).value = q.visitado
            
            ws.cell(row = cont , column=11).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=11).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=11).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=11).value = q.encontrado
            
            ws.cell(row = cont , column=12).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=12).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=12).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=12).value = q.dni_mama
            
            ws.cell(row = cont , column=13).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=13).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=13).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=13).value = q.num_cel
            
            ws.cell(row = cont , column=14).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=14).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=14).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=14).value = q.pn_reg
            
            ws.cell(row = cont , column=15).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=15).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=15).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=15).value = q.Provincia
            
            ws.cell(row = cont , column=16).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=16).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=16).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=16).value = q.Distrito
            
            ws.cell(row = cont , column=17).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=17).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=17).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=17).value = q.Red
            
            ws.cell(row = cont , column=18).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=18).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=18).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=18).value = q.MicroRed
            
            ws.cell(row = cont , column=19).alignment = Alignment(horizontal="right")
            ws.cell(row = cont , column=19).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=19).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=19).value = q.Codigo_Unico
            
            ws.cell(row = cont , column=20).alignment = Alignment(horizontal="left")
            ws.cell(row = cont , column=20).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=20).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=20).value = q.Nombre_Establecimiento
            
            ws.cell(row = cont , column=21).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=21).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=21).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=21).value = q.v_fecha1
            
            ws.cell(row = cont , column=22).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=22).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=22).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=22).value = icono_visita_1
            
            ws.cell(row = cont , column=23).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=23).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=23).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=23).value = q.v_fecha2
            
            ws.cell(row = cont , column=24).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=24).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=24).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=24).value = icono_visita_2
            
            ws.cell(row = cont , column=25).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=25).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=25).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=25).value = q.v_fecha3
            
            ws.cell(row = cont , column=26).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=26).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=26).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=26).value = icono_visita_3
            
            ws.cell(row = cont , column=27).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=27).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=27).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=27).value = q.v_fecha4
            
            ws.cell(row = cont , column=28).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=28).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=28).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=28).value = icono_visita_4
            
            ws.cell(row = cont , column=29).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=29).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=29).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=29).value = q.den
            
            ws.cell(row = cont , column=30).alignment = Alignment(horizontal="center")
            ws.cell(row = cont , column=30).border = Border(left = Side(border_style = "thin"), 
                                                                right = Side(border_style = "thin"), 
                                                                top = Side(border_style = "thin"), 
                                                                bottom = Side(border_style = "thin"))
            ws.cell(row = cont , column=30).font = Font(name = 'Calibri', size= 9)
            ws.cell(row = cont , column=30).value = q.num
            
            
            cont+=1
               
        #Respuesta de Django
        #Establecer el nombre del archivo
        nombre_archivo = "rpt_seg_visita.xlsx"
        #Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


